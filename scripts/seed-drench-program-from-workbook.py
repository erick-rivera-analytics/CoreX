from __future__ import annotations

import re
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import psycopg


ROOT = Path(r"C:\Users\paul.loja\PYPROYECTOS\dashboard_v2")
ENV_PATH = ROOT / ".env.local"
WORKBOOK_PATH = ROOT / "docs" / "drench_program_source.xlsx"

RULE_REF_TABLE = "public.field_ref_drench_program_rule_id_core_scd2"
RULE_DIM_TABLE = "public.field_dim_drench_program_rule_profile_scd2"
RULE_LINE_TABLE = "public.field_bridge_drench_program_rule_line_scd2"
PRODUCT_DIM_TABLE = "public.bodega_dim_product_profile_scd2"
PRODUCT_USAGE_TABLE = "public.bodega_bridge_product_usage_scd2"
DRENCH_ACTIVITY_ID = "FM11"

RULE_KEY_PATTERN = re.compile(r"^\s*(\d+)\s+([SP])\s+([A-Z0-9/]+)\s*$", re.IGNORECASE)
SLOT_DEFS = [
    (2, 3, 4, 5),
    (6, 7, 8, 9),
    (10, 11, 12, 13),
    (14, 15, 16, 17),
    (18, 19, 20, 21),
    (22, 23, 24, 25),
]
MANUAL_PRODUCT_CODE_MAP = {
    "HUMISTAR WG": "FB070",
    "NITRATO CALCIO": "FA007",
    "SERENADE": "PF193",
    "ALTO 100": "PF055",
    "FOLIOGOLD": "PF077",
    "FUNGAFLOR": "PF067",
    "SULFEX": "PF194",
    "FLUTRIALAQ": "PF192",
    "TACHIGAREN 36% LS": "PF117",
}


@dataclass
class ParsedLine:
    line_order: int
    application_method: str | None
    liters_per_bed: float | None
    source_product_name: str
    source_product_code: str | None
    source_unit_code: str | None
    quantity_value: float | None
    quantity_reference: str | None
    product_id: str | None


@dataclass
class ParsedRule:
    rule_id: str
    rule_code: str
    phenological_week: int
    cycle_type: str
    variety_code: str
    lines: list[ParsedLine]


def load_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    for raw in path.read_text(encoding="utf-8").splitlines():
        raw = raw.strip()
        if not raw or raw.startswith("#") or "=" not in raw:
            continue
        key, value = raw.split("=", 1)
        values[key] = value
    return values


def connect_camp():
    env = load_env(ENV_PATH)
    return psycopg.connect(
        host=env["DATABASE_HOST"],
        port=int(env["DATABASE_PORT"]),
        dbname=env.get("CAMP_DATABASE_NAME", "db_camp"),
        user=env["DATABASE_USER"],
        password=env["DATABASE_PASSWORD"],
    )


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_name(value: Any) -> str:
    return normalize_text(value).upper()


def normalize_code(value: Any) -> str:
    return normalize_text(value).upper()


def to_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number


def make_record_id() -> str:
    return str(uuid.uuid4())


def now_timestamp() -> datetime:
    return datetime.now()


def initialize_tables(cur: psycopg.Cursor[Any]) -> None:
    cur.execute(
        f"""
        create table if not exists {RULE_REF_TABLE} (
          record_id text primary key,
          rule_id text not null,
          valid_from timestamp without time zone not null,
          valid_to timestamp without time zone null,
          is_current boolean not null,
          is_valid boolean not null,
          loaded_at timestamp without time zone not null,
          run_id text not null,
          actor_id text not null,
          change_reason text not null
        )
        """
    )
    cur.execute(
        f"""
        create table if not exists {RULE_DIM_TABLE} (
          record_id text primary key,
          rule_id text not null,
          valid_from timestamp without time zone not null,
          valid_to timestamp without time zone null,
          is_current boolean not null,
          rule_code text not null,
          phenological_week integer not null,
          cycle_type text not null,
          variety_code text not null,
          activity_id text not null,
          notes text null,
          is_active boolean not null,
          is_valid boolean not null,
          loaded_at timestamp without time zone not null,
          run_id text not null,
          actor_id text not null,
          change_reason text not null
        )
        """
    )
    cur.execute(
        f"""
        create table if not exists {RULE_LINE_TABLE} (
          record_id text primary key,
          line_id text not null,
          rule_id text not null,
          valid_from timestamp without time zone not null,
          valid_to timestamp without time zone null,
          is_current boolean not null,
          line_order integer not null,
          application_method text null,
          liters_per_bed numeric(18, 6) null,
          product_id text null,
          source_product_name text null,
          source_product_code text null,
          source_unit_code text null,
          quantity_value numeric(18, 6) null,
          quantity_reference text null,
          notes text null,
          is_active boolean not null,
          is_valid boolean not null,
          loaded_at timestamp without time zone not null,
          run_id text not null,
          actor_id text not null,
          change_reason text not null
        )
        """
    )
    cur.execute(
        f"""
        create unique index if not exists field_ref_drench_program_rule_id_core_scd2_current_idx
          on {RULE_REF_TABLE} (rule_id)
          where is_current = true
        """
    )
    cur.execute(
        f"""
        create unique index if not exists field_dim_drench_program_rule_profile_scd2_current_idx
          on {RULE_DIM_TABLE} (rule_id)
          where is_current = true
        """
    )
    cur.execute(
        f"""
        create unique index if not exists field_dim_drench_program_rule_profile_scd2_key_unique_idx
          on {RULE_DIM_TABLE} (
            phenological_week,
            cycle_type,
            upper(regexp_replace(trim(variety_code), '\\s+', ' ', 'g')),
            upper(regexp_replace(trim(activity_id), '\\s+', ' ', 'g'))
          )
          where is_current = true and is_valid = true
        """
    )
    cur.execute(
        f"""
        create index if not exists field_bridge_drench_program_rule_line_scd2_current_rule_idx
          on {RULE_LINE_TABLE} (rule_id, line_order)
          where is_current = true
        """
    )


def load_code_sheet(workbook: openpyxl.Workbook) -> dict[str, dict[str, str | None]]:
    ws = workbook["CODIGOS D"]
    mapping: dict[str, dict[str, str | None]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        product_name = normalize_name(row[0])
        if not product_name:
            continue
        mapping[product_name] = {
            "code": normalize_code(row[1]) or None,
            "unit": normalize_code(row[2]) or None,
        }
    return mapping


def load_current_products(cur: psycopg.Cursor[Any]) -> tuple[dict[str, dict[str, str]], dict[str, list[str]]]:
    cur.execute(
        f"""
        select product_id, product_code, product_name
        from {PRODUCT_DIM_TABLE}
        where is_current = true
        """
    )
    by_code: dict[str, dict[str, str]] = {}
    by_name: dict[str, list[str]] = {}
    for product_id, product_code, product_name in cur.fetchall():
        code = normalize_code(product_code)
        name = normalize_name(product_name)
        by_code[code] = {
            "product_id": product_id,
            "product_code": code,
            "product_name": normalize_text(product_name),
        }
        by_name.setdefault(name, []).append(code)
    return by_code, by_name


def find_product_match(
    source_product_name: str,
    code_mapping: dict[str, dict[str, str | None]],
    products_by_code: dict[str, dict[str, str]],
    products_by_name: dict[str, list[str]],
) -> tuple[str | None, str | None, str | None]:
    normalized_name = normalize_name(source_product_name)
    source_code = code_mapping.get(normalized_name, {}).get("code")
    source_unit = code_mapping.get(normalized_name, {}).get("unit")

    if source_code and source_code in products_by_code:
        return products_by_code[source_code]["product_id"], source_code, source_unit

    manual_code = MANUAL_PRODUCT_CODE_MAP.get(normalized_name)
    if manual_code and manual_code in products_by_code:
        return products_by_code[manual_code]["product_id"], manual_code, source_unit

    name_candidates = products_by_name.get(normalized_name, [])
    if len(name_candidates) == 1:
        matched_code = name_candidates[0]
        return products_by_code[matched_code]["product_id"], matched_code, source_unit

    return None, source_code or manual_code, source_unit


def parse_rules(
    workbook: openpyxl.Workbook,
    products_by_code: dict[str, dict[str, str]],
    products_by_name: dict[str, list[str]],
) -> tuple[list[ParsedRule], dict[str, int]]:
    ws = workbook["Programa Drench"]
    headers = [normalize_text(cell.value) for cell in ws[5]]
    code_mapping = load_code_sheet(workbook)
    rules: list[ParsedRule] = []
    stats = {
        "rules": 0,
        "lines": 0,
        "matched_lines": 0,
        "unresolved_lines": 0,
    }

    for row in ws.iter_rows(min_row=6, values_only=True):
        raw_key = normalize_text(row[1] if len(row) > 1 else None)
        match = RULE_KEY_PATTERN.match(raw_key)
        if not match:
            continue

        phenological_week = int(match.group(1))
        cycle_type = match.group(2).upper()
        variety_code = match.group(3).upper()
        rule_code = f"{phenological_week} {cycle_type} {variety_code}"
        rule_id = f"drench_rule_{phenological_week:02d}_{cycle_type}_{variety_code}"
        lines: list[ParsedLine] = []

        for slot_index, (method_idx, liters_idx, product_idx, quantity_idx) in enumerate(SLOT_DEFS, start=1):
            source_product_name = normalize_text(row[product_idx] if len(row) > product_idx else None)
            if not source_product_name:
                continue

            product_id, source_product_code, source_unit_code = find_product_match(
                source_product_name,
                code_mapping,
                products_by_code,
                products_by_name,
            )
            quantity_reference = headers[quantity_idx] if quantity_idx < len(headers) else None

            lines.append(
                ParsedLine(
                    line_order=len(lines) + 1,
                    application_method=normalize_text(row[method_idx] if len(row) > method_idx else None) or None,
                    liters_per_bed=to_number(row[liters_idx] if len(row) > liters_idx else None),
                    source_product_name=source_product_name,
                    source_product_code=source_product_code,
                    source_unit_code=source_unit_code,
                    quantity_value=to_number(row[quantity_idx] if len(row) > quantity_idx else None),
                    quantity_reference=quantity_reference or None,
                    product_id=product_id,
                )
            )

            stats["lines"] += 1
            if product_id:
                stats["matched_lines"] += 1
            else:
                stats["unresolved_lines"] += 1

        rules.append(
            ParsedRule(
                rule_id=rule_id,
                rule_code=rule_code,
                phenological_week=phenological_week,
                cycle_type=cycle_type,
                variety_code=variety_code,
                lines=lines,
            )
        )
        stats["rules"] += 1

    return rules, stats


def ensure_fm11_assignments(cur: psycopg.Cursor[Any], matched_product_ids: set[str], actor_id: str, run_id: str, change_reason: str) -> int:
    inserted = 0
    for product_id in sorted(matched_product_ids):
        cur.execute(
            f"""
            select 1
            from {PRODUCT_USAGE_TABLE}
            where product_id = %s
              and activity_id = %s
              and is_current = true
            """,
            (product_id, DRENCH_ACTIVITY_ID),
        )
        if cur.fetchone():
            continue

        cur.execute(
            f"""
            select coalesce(max(branch_order), 0)
            from {PRODUCT_USAGE_TABLE}
            where product_id = %s
              and is_current = true
            """,
            (product_id,),
        )
        next_branch_order = int(cur.fetchone()[0] or 0) + 1
        now = now_timestamp()
        cur.execute(
            f"""
            insert into {PRODUCT_USAGE_TABLE} (
              record_id, product_id, valid_from, valid_to, is_current, branch_order, activity_id,
              is_valid, loaded_at, run_id, actor_id, change_reason
            )
            values (%s, %s, %s, null, true, %s, %s, true, %s, %s, %s, %s)
            """,
            (
                make_record_id(),
                product_id,
                now,
                next_branch_order,
                DRENCH_ACTIVITY_ID,
                now,
                run_id,
                actor_id,
                change_reason,
            ),
        )
        inserted += 1
    return inserted


def reseed_rules(cur: psycopg.Cursor[Any], rules: list[ParsedRule], actor_id: str, run_id: str, change_reason: str) -> None:
    now = now_timestamp()
    cur.execute(f"update {RULE_REF_TABLE} set is_current = false, valid_to = %s where is_current = true", (now,))
    cur.execute(f"update {RULE_DIM_TABLE} set is_current = false, valid_to = %s where is_current = true", (now,))
    cur.execute(f"update {RULE_LINE_TABLE} set is_current = false, valid_to = %s where is_current = true", (now,))

    for rule in rules:
        cur.execute(
            f"""
            insert into {RULE_REF_TABLE} (
              record_id, rule_id, valid_from, valid_to, is_current, is_valid, loaded_at, run_id, actor_id, change_reason
            )
            values (%s, %s, %s, null, true, true, %s, %s, %s, %s)
            """,
            (make_record_id(), rule.rule_id, now, now, run_id, actor_id, change_reason),
        )
        cur.execute(
            f"""
            insert into {RULE_DIM_TABLE} (
              record_id, rule_id, valid_from, valid_to, is_current, rule_code, phenological_week, cycle_type,
              variety_code, activity_id, notes, is_active, is_valid, loaded_at, run_id, actor_id, change_reason
            )
            values (%s, %s, %s, null, true, %s, %s, %s, %s, %s, null, true, true, %s, %s, %s, %s)
            """,
            (
                make_record_id(),
                rule.rule_id,
                now,
                rule.rule_code,
                rule.phenological_week,
                rule.cycle_type,
                rule.variety_code,
                DRENCH_ACTIVITY_ID,
                now,
                run_id,
                actor_id,
                change_reason,
            ),
        )

        for line in rule.lines:
            cur.execute(
                f"""
                insert into {RULE_LINE_TABLE} (
                  record_id, line_id, rule_id, valid_from, valid_to, is_current, line_order,
                  application_method, liters_per_bed, product_id, source_product_name, source_product_code, source_unit_code,
                  quantity_value, quantity_reference, notes, is_active, is_valid, loaded_at, run_id, actor_id, change_reason
                )
                values (%s, %s, %s, %s, null, true, %s, %s, %s, %s, %s, %s, %s, %s, %s, null, true, true, %s, %s, %s, %s)
                """,
                (
                    make_record_id(),
                    f"{rule.rule_id}_line_{line.line_order:02d}",
                    rule.rule_id,
                    now,
                    line.line_order,
                    line.application_method,
                    line.liters_per_bed,
                    line.product_id,
                    line.source_product_name,
                    line.source_product_code,
                    line.source_unit_code,
                    line.quantity_value,
                    line.quantity_reference,
                    now,
                    run_id,
                    actor_id,
                    change_reason,
                ),
            )


def main() -> None:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    actor_id = "seed_drench_program_workbook"
    run_id = f"drench_seed_{int(datetime.now().timestamp())}"
    change_reason = "INITIAL_SEED_FROM_DRENCH_WORKBOOK"

    with connect_camp() as connection:
      with connection.cursor() as cursor:
        initialize_tables(cursor)
        products_by_code, products_by_name = load_current_products(cursor)
        rules, stats = parse_rules(workbook, products_by_code, products_by_name)
        matched_product_ids = {
            line.product_id
            for rule in rules
            for line in rule.lines
            if line.product_id
        }
        inserted_assignments = ensure_fm11_assignments(cursor, matched_product_ids, actor_id, run_id, change_reason)
        reseed_rules(cursor, rules, actor_id, run_id, change_reason)
      connection.commit()

    print(
        {
            "rules": stats["rules"],
            "lines": stats["lines"],
            "matched_lines": stats["matched_lines"],
            "unresolved_lines": stats["unresolved_lines"],
            "fm11_assignments_inserted": inserted_assignments,
            "matched_products": len(matched_product_ids),
        }
    )


if __name__ == "__main__":
    main()
