from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(r"C:\Users\paul.loja\AppData\Local\Temp\CoreX_bodega_validate")
SOURCE_PATH = ROOT / "outputs" / "fumigacion_homologacion_20260525" / "fumigacion_bodega_altas_sugeridas.json"
OUTPUT_PATH = ROOT / "outputs" / "fumigacion_homologacion_20260525" / "fumigacion_revision_definitiva_codigos.xlsx"

TITLE_FILL = PatternFill(fill_type="solid", fgColor="1D4ED8")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="DBEAFE")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
HEADER_FONT = Font(color="0F172A", bold=True)


def load_rows() -> list[dict[str, object]]:
    payload = json.loads(SOURCE_PATH.read_text(encoding="utf-8"))
    return payload.get("rows", [])


def set_widths(ws, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def build_instructions_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Revision definitiva"

    ws["A1"] = "Revision definitiva de codigos Fumigacion"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:N1")

    lines = [
        ["Objetivo", "Aqui solo van los productos del Excel que no existen exactos en Bodega."],
        ["Como usarlo", "En cada fila decide si se crea un producto nuevo o si se vincula a un codigo existente de Bodega."],
        ["Decision sugerida", "Usa CREAR_NUEVO cuando realmente no exista equivalente. Usa VINCULAR_EXISTENTE cuando ya sepas a que codigo de Bodega corresponde."],
        ["Resultado esperado", "Con este archivo luego podemos automatizar altas nuevas y asignaciones de familias de fumigacion sin volver a revisar producto por producto."],
        ["Generado", datetime.now().isoformat(timespec="seconds")],
    ]

    for row_index, row in enumerate(lines, start=3):
        ws.cell(row=row_index, column=1, value=row[0]).font = HEADER_FONT
        ws.cell(row=row_index, column=2, value=row[1])

    set_widths(ws, {"A": 24, "B": 120})


def build_main_sheet(wb: Workbook, rows: list[dict[str, object]]) -> None:
    ws = wb.create_sheet("Codigos por resolver")

    headers = [
        "Producto Excel",
        "Decision usuario",
        "Codigo Bodega vincular",
        "Nombre Bodega vincular",
        "Crear codigo nuevo",
        "Nombre nuevo confirmado",
        "Unidad base confirmada",
        "Categoria Bodega confirmada",
        "Familias fumigacion requeridas",
        "Actividades fumigacion requeridas",
        "Sugerencia 1 Bodega",
        "Sugerencia 2 Bodega",
        "Programas detectados",
        "Semanas ISO ejemplo",
        "Cantidades ejemplo",
        "Observaciones",
    ]

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for record in rows:
        suggestion_1 = " | ".join(filter(None, [
            str(record.get("sugerencia_1_codigo_bodega", "")),
            str(record.get("sugerencia_1_nombre_bodega", "")),
            str(record.get("sugerencia_1_unidad", "")),
            str(record.get("sugerencia_1_categoria", "")),
        ]))
        suggestion_2 = " | ".join(filter(None, [
            str(record.get("sugerencia_2_codigo_bodega", "")),
            str(record.get("sugerencia_2_nombre_bodega", "")),
            str(record.get("sugerencia_2_unidad", "")),
            str(record.get("sugerencia_2_categoria", "")),
        ]))

        ws.append([
            record.get("producto_excel", ""),
            "",
            "",
            "",
            "",
            record.get("nuevo_product_name_propuesto", ""),
            "",
            "",
            record.get("familias_objetivo_requeridas", ""),
            record.get("actividades_objetivo_requeridas", ""),
            suggestion_1,
            suggestion_2,
            record.get("programas_detectados", ""),
            record.get("semanas_iso_ejemplo", ""),
            record.get("cantidades_ejemplo", ""),
            "",
        ])

    decision_validation = DataValidation(
        type="list",
        formula1='"VINCULAR_EXISTENTE,CREAR_NUEVO,REVISAR,NO_USAR"',
        allow_blank=True,
    )
    ws.add_data_validation(decision_validation)
    decision_validation.add(f"B2:B{ws.max_row}")

    ws.freeze_panes = "A2"
    set_widths(ws, {
        "A": 24,
        "B": 20,
        "C": 18,
        "D": 30,
        "E": 18,
        "F": 30,
        "G": 18,
        "H": 30,
        "I": 28,
        "J": 48,
        "K": 42,
        "L": 42,
        "M": 34,
        "N": 16,
        "O": 20,
        "P": 36,
    })

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    rows = load_rows()
    wb = Workbook()
    build_instructions_sheet(wb)
    build_main_sheet(wb, rows)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(str(OUTPUT_PATH))


if __name__ == "__main__":
    main()
