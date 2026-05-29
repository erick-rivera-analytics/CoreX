from __future__ import annotations

import json
import re
from difflib import SequenceMatcher
from pathlib import Path

import psycopg
from openpyxl import load_workbook


ROOT = Path(r"C:\Users\paul.loja\AppData\Local\Temp\CoreX_bodega_validate")
ENV_PATH = ROOT / ".env.local"
SEED_PATH = ROOT / "src" / "lib" / "fumigation-program-proposal-seed.ts"
REVIEW_WORKBOOK_PATH = ROOT / "outputs" / "fumigacion_homologacion_20260525" / "fumigacion_revision_definitiva_codigos.xlsx"
OUTPUT_DIR = ROOT / "outputs" / "fumigacion_homologacion_20260527"
OUTPUT_JSON_PATH = OUTPUT_DIR / "fumigacion_codigos_validacion_data.json"

FUMIGATION_ACTIVITY_IDS = {
    "FMGYP",
    "FMGYPA1",
    "FMGYPA2",
    "FMGYPM1",
    "FMGYPM2",
    "FMGYPAC2",
    "FMGYPEF",
    "03VAFIFMG",
    "03VAFIFMGL",
}

PRODUCT_NAME_ALIASES: dict[str, str] = {
    "ANTRACOL": "ANTRACOL 70% PM",
    "FITORAZ": "FITORAZ 76 PM",
    "FULMINATE": "FULMINANTE",
    "PREVICUR": "PREVICUR N",
    "CORRIDABUL": "CORRIDA BUL",
    "DANTHOTSU": "DANTOTSU 500",
    "POLYRAM": "POLYRAM 80%",
    "NIT. CA": "NITRATO DE CALCIO",
    "NIT. MG": "NITRATO DE MAGNESIO",
    "SULF DE K": "SULFATO DE POTASIO AGRICOLA",
    "NITROFOSKA 8-12-24": "NITROFOSKA AZUL 12-12-17+2",
    "METALOZATO DE CALCIO": "METALOSATO DE CALCIO",
    "TRACKING CA-B": "TRAKING CA-B",
    "DACONIL": "DACONIL 720",
    "AMISTAR": "AMISTAR TOP",
    "ORTHENE": "ORTHENE 75%",
    "POLO": "POLO 250 8C",
    "KARATE": "KARATE ZEON",
    "AVISO": "AVISO DF",
    "TIFLO": "TIFLO 42",
    "ROVRAL": "ROVRAL 50% POLVO",
    "SIVANTO PRIME 200 SL": "SIVANTO",
}


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().upper().split())


def load_env() -> dict[str, str]:
    values: dict[str, str] = {}
    for raw in ENV_PATH.read_text(encoding="utf-8").splitlines():
        raw = raw.strip()
        if not raw or raw.startswith("#") or "=" not in raw:
            continue
        key, value = raw.split("=", 1)
        values[key] = value
    return values


def parse_codes(value: object) -> list[str]:
    if value is None:
        return []
    return re.findall(r"\b[A-Z]{2}\d{3}\b", str(value).upper())


def infer_selected_code(raw_codes: list[str], observation: str) -> str:
    if not raw_codes:
        return ""

    normalized_observation = normalize_text(observation)
    if "SOLO EL PRIMERO" in normalized_observation:
        return raw_codes[0]
    if "SOLO EL SEGUNDO" in normalized_observation and len(raw_codes) >= 2:
        return raw_codes[1]
    if "SOLO EL TERCERO" in normalized_observation and len(raw_codes) >= 3:
        return raw_codes[2]
    return raw_codes[0]


def load_seed_names() -> dict[str, dict[str, object]]:
    text = SEED_PATH.read_text(encoding="utf-8")
    names = re.findall(r'"productName":"([^"]+)"', text)
    seed_map: dict[str, dict[str, object]] = {}

    for source_name in names:
        normalized_source = normalize_text(source_name)
        canonical_target = PRODUCT_NAME_ALIASES.get(normalized_source, source_name)
        canonical_key = normalize_text(canonical_target)
        entry = seed_map.setdefault(
            canonical_key,
            {
                "canonical_name_hint": canonical_target,
                "seed_source_names": [],
                "seed_occurrences": 0,
            },
        )
        seed_source_names: list[str] = entry["seed_source_names"]  # type: ignore[assignment]
        if source_name not in seed_source_names:
            seed_source_names.append(source_name)
        entry["seed_occurrences"] = int(entry["seed_occurrences"]) + 1  # type: ignore[arg-type]

    return seed_map


def load_review_rows() -> dict[str, dict[str, object]]:
    workbook = load_workbook(REVIEW_WORKBOOK_PATH, data_only=True)
    sheet = workbook["Codigos por resolver"]
    headers = [cell.value for cell in sheet[1]]
    lookup: dict[str, dict[str, object]] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        record = {str(headers[index]): row[index] for index in range(len(headers))}
        lookup[normalize_text(row[0])] = record

    return lookup


def fetch_bodega_products() -> dict[str, dict[str, object]]:
    env = load_env()
    conn = psycopg.connect(
        host=env["DATABASE_HOST"],
        port=int(env["DATABASE_PORT"]),
        dbname=env.get("BODEGA_DATABASE_NAME", "db_storageroom"),
        user=env["DATABASE_USER"],
        password=env["DATABASE_PASSWORD"],
    )

    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                select
                  p.product_id,
                  p.product_code,
                  p.product_name,
                  units.unit_code,
                  units.unit_name,
                  categories.category_name,
                  coalesce(array_agg(distinct upper(trim(u.activity_id))) filter (where u.activity_id is not null), '{}') as activity_ids
                from public.sr_dim_product_profile_scd2 p
                inner join public.sr_ref_product_id_core_scd2 ref
                  on ref.product_id = p.product_id
                 and ref.is_current = true
                inner join public.sr_dim_unit_profile_scd2 units
                  on units.unit_id = p.base_unit_id
                 and units.is_current = true
                left join (
                  select category_id, category_name
                  from public.sr_dim_category_profile_scd2
                  where is_current = true
                ) categories
                  on categories.category_id = p.category_id
                left join public.sr_bridge_product_usage_scd2 u
                  on u.product_id = p.product_id
                 and u.is_current = true
                 and u.is_valid = true
                where p.is_current = true
                  and p.is_valid = true
                group by
                  p.product_id,
                  p.product_code,
                  p.product_name,
                  units.unit_code,
                  units.unit_name,
                  categories.category_name
                order by upper(trim(p.product_name)), upper(trim(p.product_code))
                """
            )

            rows = cur.fetchall()
    finally:
        conn.close()

    by_name: dict[str, dict[str, object]] = {}
    for product_id, product_code, product_name, unit_code, unit_name, category_name, activity_ids in rows:
        by_name[normalize_text(product_name)] = {
            "product_id": product_id,
            "product_code": product_code,
            "product_name": product_name,
            "unit_code": unit_code,
            "unit_name": unit_name,
            "category_name": category_name or "",
            "activity_ids": sorted(activity_ids or []),
            "has_fumigation_family": bool(set(activity_ids or []) & FUMIGATION_ACTIVITY_IDS),
        }

    return by_name


def tokenize_similarity_name(value: str) -> list[str]:
    cleaned = re.sub(r"[^A-Z0-9]+", " ", normalize_text(value))
    stopwords = {
        "DE",
        "DEL",
        "LA",
        "EL",
        "Y",
        "CON",
        "PARA",
        "TOP",
        "DF",
        "PM",
        "SL",
        "SC",
        "WG",
        "OD",
        "EC",
        "GR",
        "CC",
        "POLVO",
        "PRIME",
        "FOLIAR",
        "AZUL",
    }
    return [token for token in cleaned.split() if len(token) >= 3 and token not in stopwords]


def compact_similarity_name(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", normalize_text(value))


def find_similar_bodega_candidates(
    canonical_name: str,
    seed_source_names: list[str],
    bodega_by_name: dict[str, dict[str, object]],
) -> list[dict[str, object]]:
    search_terms = list(dict.fromkeys([canonical_name, *seed_source_names]))
    search_token_sets = [set(tokenize_similarity_name(term)) for term in search_terms if tokenize_similarity_name(term)]
    search_compacts = [compact_similarity_name(term) for term in search_terms if compact_similarity_name(term)]

    candidates: list[dict[str, object]] = []
    for product in bodega_by_name.values():
        product_name = str(product["product_name"])
        product_tokens = set(tokenize_similarity_name(product_name))
        product_compact = compact_similarity_name(product_name)
        best_score = 0.0
        matched = False

        for token_set, search_compact in zip(search_token_sets, search_compacts, strict=False):
            overlap = len(token_set & product_tokens)
            ratio = SequenceMatcher(None, search_compact, product_compact).ratio()
            first_token_match = bool(token_set and product_tokens and next(iter(token_set)) == next(iter(product_tokens)))

            if search_compact and (search_compact in product_compact or product_compact in search_compact):
                matched = True
                best_score = max(best_score, 1.0)
                continue

            if overlap >= 1 and (ratio >= 0.55 or first_token_match):
                matched = True
                best_score = max(best_score, ratio + (0.1 * overlap))
                continue

            if ratio >= 0.74:
                matched = True
                best_score = max(best_score, ratio)

        if matched:
            candidate = dict(product)
            candidate["similarity_score"] = round(best_score, 3)
            candidates.append(candidate)

    candidates.sort(
        key=lambda item: (
            -float(item["similarity_score"]),
            normalize_text(item["product_name"]),
            normalize_text(item["product_code"]),
        )
    )
    return candidates


def compute_rows() -> dict[str, object]:
    seed_map = load_seed_names()
    review_lookup = load_review_rows()
    bodega_by_name = fetch_bodega_products()

    rows: list[dict[str, object]] = []
    explicit_multi_code_rows: list[dict[str, object]] = []

    for review_key, review_row in sorted(review_lookup.items()):
        raw_codes = parse_codes(review_row.get("Codigo Bodega vincular"))
        if len(raw_codes) <= 1:
            continue

        selected_code = infer_selected_code(raw_codes, str(review_row.get("Observaciones") or ""))
        unused_codes = [code for code in raw_codes if code != selected_code]
        selected_bodega = next(
            (
                product
                for product in bodega_by_name.values()
                if normalize_text(product["product_code"]) == normalize_text(selected_code)
            ),
            None,
        )

        explicit_multi_code_rows.append(
            {
                "producto_excel": review_key,
                "decision_usuario": str(review_row.get("Decision usuario") or ""),
                "codigo_bodega_vincular_raw": str(review_row.get("Codigo Bodega vincular") or ""),
                "codigo_seleccionado": selected_code,
                "nombre_bodega_seleccionado": str(selected_bodega["product_name"]) if selected_bodega else "",
                "codigos_no_usados": ", ".join(unused_codes),
                "sugerencia_1_bodega": str(review_row.get("Sugerencia 1 Bodega") or ""),
                "sugerencia_2_bodega": str(review_row.get("Sugerencia 2 Bodega") or ""),
                "familias_requeridas": str(review_row.get("Familias fumigacion requeridas") or ""),
                "actividades_requeridas": str(review_row.get("Actividades fumigacion requeridas") or ""),
                "programas_detectados": str(review_row.get("Programas detectados") or ""),
                "semanas_iso_ejemplo": str(review_row.get("Semanas ISO ejemplo") or ""),
                "cantidades_ejemplo": str(review_row.get("Cantidades ejemplo") or ""),
                "observaciones": str(review_row.get("Observaciones") or ""),
            }
        )

    for canonical_key in sorted(seed_map.keys()):
        seed_entry = seed_map[canonical_key]
        review_row = review_lookup.get(canonical_key)
        bodega_product = bodega_by_name.get(canonical_key)
        similar_candidates = find_similar_bodega_candidates(
            str(seed_entry["canonical_name_hint"]),
            list(seed_entry["seed_source_names"]),  # type: ignore[arg-type]
            bodega_by_name,
        )

        considered_codes: list[str] = []
        if review_row is not None:
            considered_codes.extend(parse_codes(review_row.get("Codigo Bodega vincular")))
            considered_codes.extend(parse_codes(review_row.get("Sugerencia 1 Bodega")))
            considered_codes.extend(parse_codes(review_row.get("Sugerencia 2 Bodega")))
        considered_codes = list(dict.fromkeys(considered_codes))

        used_code = str(bodega_product["product_code"]) if bodega_product else ""
        used_name = str(bodega_product["product_name"]) if bodega_product else ""
        unused_codes = [code for code in considered_codes if code != used_code]
        status = "RESUELTO" if bodega_product and bodega_product.get("has_fumigation_family") else "PENDIENTE"

        rows.append(
            {
                "producto_programacion": used_name or str(seed_entry["canonical_name_hint"]),
                "producto_programacion_key": canonical_key,
                "nombres_semilla": ", ".join(sorted(seed_entry["seed_source_names"])),  # type: ignore[arg-type]
                "apariciones_semilla": int(seed_entry["seed_occurrences"]),
                "estado": status,
                "codigo_usado": used_code,
                "nombre_bodega_usado": used_name,
                "unidad_codigo": str(bodega_product["unit_code"]) if bodega_product else "",
                "unidad_nombre": str(bodega_product["unit_name"]) if bodega_product else "",
                "categoria_bodega": str(bodega_product["category_name"]) if bodega_product else "",
                "decision_usuario": str(review_row.get("Decision usuario") or "") if review_row else "",
                "codigo_bodega_vincular_raw": str(review_row.get("Codigo Bodega vincular") or "") if review_row else "",
                "codigos_considerados": ", ".join(considered_codes),
                "codigos_no_usados": ", ".join(unused_codes),
                "sugerencia_1_bodega": str(review_row.get("Sugerencia 1 Bodega") or "") if review_row else "",
                "sugerencia_2_bodega": str(review_row.get("Sugerencia 2 Bodega") or "") if review_row else "",
                "familias_requeridas": str(review_row.get("Familias fumigacion requeridas") or "") if review_row else "",
                "actividades_requeridas": str(review_row.get("Actividades fumigacion requeridas") or "") if review_row else "",
                "programas_detectados": str(review_row.get("Programas detectados") or "") if review_row else "",
                "semanas_iso_ejemplo": str(review_row.get("Semanas ISO ejemplo") or "") if review_row else "",
                "cantidades_ejemplo": str(review_row.get("Cantidades ejemplo") or "") if review_row else "",
                "observaciones": str(review_row.get("Observaciones") or "") if review_row else "",
                "similar_bodega_candidates": [
                    {
                        "product_code": str(candidate["product_code"]),
                        "product_name": str(candidate["product_name"]),
                        "unit_code": str(candidate["unit_code"]),
                        "category_name": str(candidate["category_name"]),
                        "has_fumigation_family": bool(candidate["has_fumigation_family"]),
                        "similarity_score": float(candidate["similarity_score"]),
                    }
                    for candidate in similar_candidates
                ],
            }
        )

    pending_rows = [row for row in rows if row["estado"] == "PENDIENTE"]
    multi_code_rows = explicit_multi_code_rows
    similar_name_rows = [
        {
            **row,
            "similar_bodega_candidates_text": " | ".join(
                f"{candidate['product_code']} · {candidate['product_name']} · {candidate['unit_code']} · "
                f"{'SI' if candidate['has_fumigation_family'] else 'NO'} · score {candidate['similarity_score']}"
                for candidate in row["similar_bodega_candidates"]
            ),
        }
        for row in rows
        if len(row["similar_bodega_candidates"]) > 1
    ]

    return {
        "generated_from_seed": SEED_PATH.name,
        "generated_from_review_workbook": REVIEW_WORKBOOK_PATH.name,
        "total_productos_programacion": len(rows),
        "total_resueltos": len(rows) - len(pending_rows),
        "total_pendientes": len(pending_rows),
        "total_multicodigo": len(multi_code_rows),
        "total_similares_bodega": len(similar_name_rows),
        "rows": rows,
        "pending_rows": pending_rows,
        "multi_code_rows": multi_code_rows,
        "similar_name_rows": similar_name_rows,
    }


def main() -> None:
    payload = compute_rows()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(OUTPUT_JSON_PATH)


if __name__ == "__main__":
    main()
