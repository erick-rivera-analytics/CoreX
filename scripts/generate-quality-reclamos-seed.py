from __future__ import annotations

import hashlib
import re
import unicodedata
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = Path(r"C:\Users\paul.loja\Downloads\Selecionables_APP_Reclamos.xlsx")
OUTPUT_PATH = ROOT / "sql" / "db_quality_reclamos_seed.sql"
RUN_ID = "seed_reclamos_selectables_xlsx_20260511_v2"
SEED_TS = "2026-05-11 00:00:00"


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip())


def slug_base(value: str) -> str:
    text = normalize_space(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", "_", text).strip("_")
    return text or "ITEM"


def make_seed_id(prefix: str, key: str) -> str:
    digest = hashlib.sha1(key.encode("utf-8")).hexdigest()[:16]
    return f"{prefix}_{digest}"


def make_code(prefix: str, value: str) -> str:
    base = slug_base(value)
    if len(base) > 48:
      base = base[:48].rstrip("_")
    return f"{prefix}_{base}"


def sql_text(value: str | None) -> str:
    if value is None:
        return "null"
    return "'" + value.replace("'", "''") + "'"


def ordered_unique(values: list[str]) -> list[str]:
    return list(OrderedDict((value, None) for value in values))


def read_unique_column(ws, header_name: str) -> list[str]:
    headers = [ws.cell(row=1, column=idx).value for idx in range(1, ws.max_column + 1)]
    if header_name not in headers:
        return []
    col_idx = headers.index(header_name) + 1
    values: list[str] = []
    for row_idx in range(2, ws.max_row + 1):
        raw = ws.cell(row=row_idx, column=col_idx).value
        if raw is None:
            continue
        text = normalize_space(str(raw))
        if text:
            values.append(text)
    return ordered_unique(values)


def build_simple_seed_rows(items: list[str], code_prefix: str, entity_prefix: str, label: str):
    rows: list[dict[str, str]] = []
    for item in items:
        entity_id = make_seed_id(f"seed_{entity_prefix}", item)
        rows.append(
            {
                "ref_record_id": make_seed_id(f"seed_ref_{entity_prefix}", item),
                "dim_record_id": make_seed_id(f"seed_dim_{entity_prefix}", item),
                "entity_id": entity_id,
                "entity_code": make_code(code_prefix, item),
                "entity_name": item,
                "entity_description": f"Semilla inicial de {label} desde Excel",
            }
        )
    return rows


def render_values(rows: list[tuple[str, ...]]) -> str:
    return ",\n".join(f"  ({', '.join(parts)})" for parts in rows)


def render_simple_master_sql(ref_table: str, dim_table: str, rows: list[dict[str, str]]) -> str:
    ref_values = []
    dim_values = []
    for row in rows:
        ref_values.append(
            (
                sql_text(row["ref_record_id"]),
                sql_text(row["entity_id"]),
                f"timestamp {sql_text(SEED_TS)}",
                "null::timestamp",
                "true",
                "true",
                "now()",
                sql_text(RUN_ID),
                sql_text("corex_seed"),
                sql_text(row["entity_description"]),
            )
        )
        dim_values.append(
            (
                sql_text(row["dim_record_id"]),
                sql_text(row["entity_id"]),
                f"timestamp {sql_text(SEED_TS)}",
                "null::timestamp",
                "true",
                sql_text(row["entity_code"]),
                sql_text(row["entity_name"]),
                sql_text(row["entity_description"]),
                "null",
                "null",
                "true",
                "true",
                "now()",
                sql_text(RUN_ID),
                sql_text("corex_seed"),
                sql_text(row["entity_description"]),
            )
        )

    return f"""
insert into {ref_table} (
  record_id, entity_id, valid_from, valid_to, is_current, is_valid, loaded_at, run_id, actor_id, change_reason
)
select *
from (
values
{render_values(ref_values)}
) as seed_rows(record_id, entity_id, valid_from, valid_to, is_current, is_valid, loaded_at, run_id, actor_id, change_reason)
where not exists (
  select 1 from {ref_table} existing where existing.record_id = seed_rows.record_id
);

insert into {dim_table} (
  record_id, entity_id, valid_from, valid_to, is_current, entity_code, entity_name, entity_description,
  external_ref_code, contact_email, is_active, is_valid, loaded_at, run_id, actor_id, change_reason
)
select *
from (
values
{render_values(dim_values)}
) as seed_rows(record_id, entity_id, valid_from, valid_to, is_current, entity_code, entity_name, entity_description, external_ref_code, contact_email, is_active, is_valid, loaded_at, run_id, actor_id, change_reason)
where not exists (
  select 1 from {dim_table} existing where existing.record_id = seed_rows.record_id
);
""".strip()


def build_problem_rows(scope: str, families: list[str], problems: list[str]) -> list[dict[str, str | int | None]]:
    rows: list[dict[str, str | int | None]] = []
    scope_prefix = "Q" if scope == "quality" else "C"

    for index, family in enumerate(families, start=1):
        problem_id = make_seed_id(f"seed_claim_family_{scope}", family)
        rows.append(
            {
                "ref_record_id": make_seed_id(f"seed_ref_claim_family_{scope}", family),
                "dim_record_id": make_seed_id(f"seed_dim_claim_family_{scope}", family),
                "problem_id": problem_id,
                "problem_code": make_code(f"{scope_prefix}FAM", family),
                "problem_name": family,
                "problem_level": "family",
                "problem_scope": scope,
                "parent_problem_id": None,
                "sort_order": index * 10,
                "problem_description": "Semilla inicial de tipos de problema desde Excel",
            }
        )

    for index, problem in enumerate(problems, start=1):
        problem_id = make_seed_id(f"seed_claim_problem_{scope}", problem)
        rows.append(
            {
                "ref_record_id": make_seed_id(f"seed_ref_claim_problem_{scope}", problem),
                "dim_record_id": make_seed_id(f"seed_dim_claim_problem_{scope}", problem),
                "problem_id": problem_id,
                "problem_code": make_code(f"{scope_prefix}PRB", problem),
                "problem_name": problem,
                "problem_level": "subfamily",
                "problem_scope": scope,
                "parent_problem_id": None,
                "sort_order": index * 10,
                "problem_description": "Semilla inicial de problemas de reclamo desde Excel sin relacion de familia",
            }
        )

    return rows


def render_problem_sql(rows: list[dict[str, str | int | None]]) -> str:
    ref_values = []
    dim_values = []

    for row in rows:
        ref_values.append(
            (
                sql_text(str(row["ref_record_id"])),
                sql_text(str(row["problem_id"])),
                f"timestamp {sql_text(SEED_TS)}",
                "null::timestamp",
                "true",
                "true",
                "now()",
                sql_text(RUN_ID),
                sql_text("corex_seed"),
                sql_text(str(row["problem_description"])),
            )
        )
        dim_values.append(
            (
                sql_text(str(row["dim_record_id"])),
                sql_text(str(row["problem_id"])),
                f"timestamp {sql_text(SEED_TS)}",
                "null::timestamp",
                "true",
                sql_text(str(row["problem_code"])),
                sql_text(str(row["problem_name"])),
                sql_text(str(row["problem_level"])),
                sql_text(str(row["problem_scope"])),
                sql_text(row["parent_problem_id"]),
                str(row["sort_order"]),
                sql_text(str(row["problem_description"])),
                "true",
                "true",
                "now()",
                sql_text(RUN_ID),
                sql_text("corex_seed"),
                sql_text(str(row["problem_description"])),
            )
        )

    return f"""
insert into public.sls_ref_claim_problem_id_core_scd2 (
  record_id, problem_id, valid_from, valid_to, is_current, is_valid, loaded_at, run_id, actor_id, change_reason
)
select *
from (
values
{render_values(ref_values)}
) as seed_rows(record_id, problem_id, valid_from, valid_to, is_current, is_valid, loaded_at, run_id, actor_id, change_reason)
where not exists (
  select 1 from public.sls_ref_claim_problem_id_core_scd2 existing where existing.record_id = seed_rows.record_id
);

insert into public.sls_dim_claim_problem_profile_scd2 (
  record_id, problem_id, valid_from, valid_to, is_current, problem_code, problem_name, problem_level,
  problem_scope, parent_problem_id, sort_order, problem_description, is_active, is_valid,
  loaded_at, run_id, actor_id, change_reason
)
select *
from (
values
{render_values(dim_values)}
) as seed_rows(record_id, problem_id, valid_from, valid_to, is_current, problem_code, problem_name, problem_level, problem_scope, parent_problem_id, sort_order, problem_description, is_active, is_valid, loaded_at, run_id, actor_id, change_reason)
where not exists (
  select 1 from public.sls_dim_claim_problem_profile_scd2 existing where existing.record_id = seed_rows.record_id
);
""".strip()


def main() -> None:
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws_quality = wb["Calidad"]
    ws_commercial = wb["Comercial"]

    customers = ordered_unique(
        read_unique_column(ws_quality, "Clientes") + read_unique_column(ws_commercial, "Clientes")
    )
    commercializers = ordered_unique(
        read_unique_column(ws_quality, "Comercializadora") + read_unique_column(ws_commercial, "Comercializadora")
    )
    executives = ordered_unique(
        read_unique_column(ws_quality, "Ejecutivo de Cuenta") + read_unique_column(ws_commercial, "Ejecutivo de Cuenta")
    )

    quality_families = read_unique_column(ws_quality, "Tipo de Problema")
    quality_problems = read_unique_column(ws_quality, "Problema")
    commercial_families = read_unique_column(ws_commercial, "Tipo de Problema")
    commercial_problems = read_unique_column(ws_commercial, "Problema")

    customer_rows = build_simple_seed_rows(customers, "CUST", "customer", "clientes")
    commercializer_rows = build_simple_seed_rows(commercializers, "COM", "commercializer", "comercializadoras")
    executive_rows = build_simple_seed_rows(executives, "EXEC", "executive", "ejecutivos de cuenta")

    problem_rows = build_problem_rows("quality", quality_families, quality_problems) + build_problem_rows(
        "commercial",
        commercial_families,
        commercial_problems,
    )

    output = f"""-- =========================================================================
-- db_quality_reclamos_seed.sql - Semilla inicial desde Selecionables_APP_Reclamos.xlsx
-- Aplicar contra: db_calidad
-- Nota: los problemas quedan con parent_problem_id = null hasta recibir el mapeo final de familias.
-- =========================================================================

begin;

delete from public.sls_dim_claim_problem_profile_scd2
where run_id like 'seed_reclamos_selectables_xlsx_%';

delete from public.sls_ref_claim_problem_id_core_scd2
where run_id like 'seed_reclamos_selectables_xlsx_%';

delete from public.sls_dim_account_executive_profile_scd2
where run_id like 'seed_reclamos_selectables_xlsx_%';

delete from public.sls_ref_account_executive_id_core_scd2
where run_id like 'seed_reclamos_selectables_xlsx_%';

delete from public.sls_dim_commercializer_profile_scd2
where run_id like 'seed_reclamos_selectables_xlsx_%';

delete from public.sls_ref_commercializer_id_core_scd2
where run_id like 'seed_reclamos_selectables_xlsx_%';

delete from public.sls_dim_customer_profile_scd2
where run_id like 'seed_reclamos_selectables_xlsx_%';

delete from public.sls_ref_customer_id_core_scd2
where run_id like 'seed_reclamos_selectables_xlsx_%';

{render_simple_master_sql('public.sls_ref_customer_id_core_scd2', 'public.sls_dim_customer_profile_scd2', customer_rows)}

{render_simple_master_sql('public.sls_ref_commercializer_id_core_scd2', 'public.sls_dim_commercializer_profile_scd2', commercializer_rows)}

{render_simple_master_sql('public.sls_ref_account_executive_id_core_scd2', 'public.sls_dim_account_executive_profile_scd2', executive_rows)}

{render_problem_sql(problem_rows)}

commit;
"""

    OUTPUT_PATH.write_text(output, encoding="utf-8")
    print(f"Generado: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
