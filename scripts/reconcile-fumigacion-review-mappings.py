from __future__ import annotations

import json
import re
import uuid
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import psycopg
from openpyxl import load_workbook


ROOT = Path(r"C:\Users\paul.loja\AppData\Local\Temp\CoreX_bodega_validate")
ENV_PATH = ROOT / ".env.local"
WORKBOOK_PATH = ROOT / "outputs" / "fumigacion_homologacion_20260525" / "fumigacion_revision_definitiva_codigos.xlsx"
OUTPUT_PATH = ROOT / "outputs" / "fumigacion_homologacion_20260525" / "fumigacion_review_reconcile_results.json"

PRODUCT_REF_TABLE = "public.sr_ref_product_id_core_scd2"
PRODUCT_DIM_TABLE = "public.sr_dim_product_profile_scd2"
PRODUCT_USAGE_TABLE = "public.sr_bridge_product_usage_scd2"

ACTOR_ID = "codex_fumigation_review_reconcile"
CHANGE_REASON = "RECONCILE_FUMIGATION_REVIEW_MAPPING"
AUTOMATION_ACTORS = {
    "codex_fumigation_review_mapping",
    "codex_fumigation_review_reconcile",
}
AUTOMATION_REASONS = {
    "APPLY_FUMIGATION_REVIEW_CODE_MAPPING",
    "RECONCILE_FUMIGATION_REVIEW_MAPPING",
}


@dataclass
class CurrentProduct:
    product_id: str
    product_code: str
    product_name: str
    product_description: str | None
    base_unit_id: str
    category_id: str
    active_component_mode: str
    active_component_name: str | None
    is_active: bool
    valid_from: datetime


def load_env() -> dict[str, str]:
    values: dict[str, str] = {}
    for raw in ENV_PATH.read_text(encoding="utf-8").splitlines():
        raw = raw.strip()
        if not raw or raw.startswith("#") or "=" not in raw:
            continue
        key, value = raw.split("=", 1)
        values[key] = value
    return values


def make_record_id() -> str:
    return str(uuid.uuid4())


def make_run_id() -> str:
    return f"bodega_product_update_{datetime.now().strftime('%Y%m%d%H%M%S')}"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().upper().split())


def parse_activity_list(value: object) -> list[str]:
    raw = normalize_text(value)
    if not raw:
        return []
    return [part.strip() for part in raw.split(",") if part.strip()]


def parse_codes(value: object) -> list[str]:
    raw = normalize_text(value)
    if not raw:
        return []
    return re.findall(r"[A-Z]{2}\d{3}", raw)


def choose_codes(codes: list[str], observation: str) -> tuple[list[str], list[str]]:
    if len(codes) <= 1:
      return codes, []

    observation = normalize_text(observation)
    if "SOLO EL SEGUNDO" in observation and len(codes) >= 2:
        return [codes[1]], [code for index, code in enumerate(codes) if index != 1]
    if "SOLO EL PRIMERO" in observation and len(codes) >= 1:
        return [codes[0]], codes[1:]
    if "SOLO EL TERCERO" in observation and len(codes) >= 3:
        return [codes[2]], [code for index, code in enumerate(codes) if index != 2]

    return codes, []


def load_review_targets() -> tuple[dict[str, set[str]], set[str], list[dict[str, object]]]:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Codigos por resolver"]
    headers = [cell.value for cell in ws[1]]

    desired_code_activities: dict[str, set[str]] = defaultdict(set)
    undesired_codes: set[str] = set()
    pending_rows: list[dict[str, object]] = []

    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        product_name = normalize_text(row.get("Producto Excel"))
        decision = normalize_text(row.get("Decision usuario"))
        observation = str(row.get("Observaciones") or "")
        activities = parse_activity_list(row.get("Actividades fumigacion requeridas"))
        codes = parse_codes(row.get("Codigo Bodega vincular"))

        if decision != "VINCULAR_EXISTENTE":
            pending_rows.append({
                "producto_excel": product_name,
                "decision_usuario": decision,
                "codigo_bodega_vincular": row.get("Codigo Bodega vincular"),
                "observaciones": observation,
            })
            continue

        if not codes:
            pending_rows.append({
                "producto_excel": product_name,
                "decision_usuario": decision,
                "codigo_bodega_vincular": row.get("Codigo Bodega vincular"),
                "observaciones": observation,
            })
            continue

        selected_codes, dropped_codes = choose_codes(codes, observation)
        for code in selected_codes:
            desired_code_activities[code].update(activities)
        for code in dropped_codes:
            undesired_codes.add(code)

    return desired_code_activities, undesired_codes, pending_rows


def fetch_current_products(conn: psycopg.Connection) -> tuple[dict[str, CurrentProduct], dict[str, list[str]]]:
    with conn.cursor() as cur:
        cur.execute(
            f"""
            select
              p.product_id,
              upper(trim(p.product_code)) as product_code,
              upper(trim(p.product_name)) as product_name,
              p.product_description,
              p.base_unit_id,
              p.category_id,
              p.active_component_mode,
              p.active_component_name,
              coalesce(p.is_active, true) as is_active,
              p.valid_from
            from {PRODUCT_DIM_TABLE} p
            where p.is_current = true
              and p.is_valid = true
            """
        )
        product_rows = cur.fetchall()

        cur.execute(
            f"""
            select
              u.product_id,
              upper(trim(u.activity_id)) as activity_id
            from {PRODUCT_USAGE_TABLE} u
            where u.is_current = true
              and u.is_valid = true
            order by u.product_id, coalesce(u.branch_order, 0), upper(trim(u.activity_id))
            """
        )
        usage_rows = cur.fetchall()

    by_code: dict[str, CurrentProduct] = {}
    for row in product_rows:
        by_code[row[1]] = CurrentProduct(
            product_id=row[0],
            product_code=row[1],
            product_name=row[2],
            product_description=row[3],
            base_unit_id=row[4],
            category_id=row[5],
            active_component_mode=row[6],
            active_component_name=row[7],
            is_active=bool(row[8]),
            valid_from=row[9],
        )

    usage_by_product: dict[str, list[str]] = defaultdict(list)
    for product_id, activity_id in usage_rows:
        if activity_id not in usage_by_product[product_id]:
            usage_by_product[product_id].append(activity_id)

    return by_code, usage_by_product


def fetch_previous_activities(conn: psycopg.Connection, product_id: str, current_valid_from: datetime) -> list[str]:
    with conn.cursor() as cur:
        cur.execute(
            f"""
            select max(valid_from)
            from {PRODUCT_USAGE_TABLE}
            where product_id = %s
              and valid_from < %s
              and coalesce(actor_id, '') <> all(%s)
              and coalesce(change_reason, '') <> all(%s)
            """,
            (product_id, current_valid_from, list(AUTOMATION_ACTORS), list(AUTOMATION_REASONS)),
        )
        prev_valid_from = cur.fetchone()[0]
        if prev_valid_from is None:
            return []

        cur.execute(
            f"""
            select distinct upper(trim(activity_id)) as activity_id
            from {PRODUCT_USAGE_TABLE}
            where product_id = %s
              and valid_from = %s
              and is_valid = true
            order by 1
            """,
            (product_id, prev_valid_from),
        )
        return [row[0] for row in cur.fetchall()]


def update_product_assignments(conn: psycopg.Connection, product: CurrentProduct, final_activities: list[str]) -> None:
    now = datetime.now()
    run_id = make_run_id()

    with conn.cursor() as cur:
        cur.execute(
            f"update {PRODUCT_REF_TABLE} set is_current = false, valid_to = %s where product_id = %s and is_current = true",
            (now, product.product_id),
        )
        cur.execute(
            f"update {PRODUCT_DIM_TABLE} set is_current = false, valid_to = %s where product_id = %s and is_current = true",
            (now, product.product_id),
        )
        cur.execute(
            f"update {PRODUCT_USAGE_TABLE} set is_current = false, valid_to = %s where product_id = %s and is_current = true",
            (now, product.product_id),
        )

        cur.execute(
            f"""
            insert into {PRODUCT_REF_TABLE} (
              record_id, product_id, valid_from, valid_to, is_current, is_valid, loaded_at, run_id, actor_id, change_reason
            ) values (%s, %s, %s, null, true, true, %s, %s, %s, %s)
            """,
            (make_record_id(), product.product_id, now, now, run_id, ACTOR_ID, CHANGE_REASON),
        )

        cur.execute(
            f"""
            insert into {PRODUCT_DIM_TABLE} (
              record_id, product_id, valid_from, valid_to, is_current, product_code, product_name,
              product_description, base_unit_id, category_id, active_component_mode, active_component_name, is_active,
              is_valid, loaded_at, run_id, actor_id, change_reason
            ) values (%s, %s, %s, null, true, %s, %s, %s, %s, %s, %s, %s, %s, true, %s, %s, %s, %s)
            """,
            (
                make_record_id(),
                product.product_id,
                now,
                product.product_code,
                product.product_name,
                product.product_description,
                product.base_unit_id,
                product.category_id,
                product.active_component_mode,
                product.active_component_name,
                product.is_active,
                now,
                run_id,
                ACTOR_ID,
                CHANGE_REASON,
            ),
        )

        for index, activity_id in enumerate(final_activities, start=1):
            cur.execute(
                f"""
                insert into {PRODUCT_USAGE_TABLE} (
                  record_id, product_id, valid_from, valid_to, is_current, branch_order, activity_id,
                  is_valid, loaded_at, run_id, actor_id, change_reason
                ) values (%s, %s, %s, null, true, %s, %s, true, %s, %s, %s, %s)
                """,
                (
                    make_record_id(),
                    product.product_id,
                    now,
                    index,
                    activity_id,
                    now,
                    run_id,
                    ACTOR_ID,
                    CHANGE_REASON,
                ),
            )


def main() -> None:
    desired_code_activities, undesired_codes, pending_rows = load_review_targets()
    env = load_env()
    conn = psycopg.connect(
        host=env["DATABASE_HOST"],
        port=int(env["DATABASE_PORT"]),
        dbname=env.get("BODEGA_DATABASE_NAME", "db_storageroom"),
        user=env["DATABASE_USER"],
        password=env["DATABASE_PASSWORD"],
    )
    conn.autocommit = True

    results = {
        "generated_at": datetime.now().isoformat(),
        "desired_codes_applied": [],
        "undesired_codes_reverted": [],
        "pending_rows": pending_rows,
        "codes_not_found": [],
    }

    try:
        products_by_code, usage_by_product = fetch_current_products(conn)

        for code, required_activities in sorted(desired_code_activities.items()):
            product = products_by_code.get(code)
            if not product:
                results["codes_not_found"].append({"product_code": code, "mode": "desired"})
                continue

            current_activities = usage_by_product.get(product.product_id, [])
            final_activities = []
            for activity in current_activities + sorted(required_activities):
                if activity not in final_activities:
                    final_activities.append(activity)

            if final_activities == current_activities:
                continue

            with conn.transaction():
                update_product_assignments(conn, product, final_activities)

            results["desired_codes_applied"].append({
                "product_code": code,
                "product_name": product.product_name,
                "final_activities": final_activities,
            })

        products_by_code, usage_by_product = fetch_current_products(conn)
        for code in sorted(undesired_codes):
            product = products_by_code.get(code)
            if not product:
                results["codes_not_found"].append({"product_code": code, "mode": "undesired"})
                continue

            previous_activities = fetch_previous_activities(conn, product.product_id, product.valid_from)
            current_activities = usage_by_product.get(product.product_id, [])

            if current_activities == previous_activities:
                continue

            with conn.transaction():
                update_product_assignments(conn, product, previous_activities)

            results["undesired_codes_reverted"].append({
                "product_code": code,
                "product_name": product.product_name,
                "restored_activities": previous_activities,
            })

        OUTPUT_PATH.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps({
            "desired_codes_applied": len(results["desired_codes_applied"]),
            "undesired_codes_reverted": len(results["undesired_codes_reverted"]),
            "pending_rows": len(results["pending_rows"]),
            "codes_not_found": len(results["codes_not_found"]),
            "output": str(OUTPUT_PATH),
        }, ensure_ascii=False))
    finally:
        conn.close()


if __name__ == "__main__":
    main()
