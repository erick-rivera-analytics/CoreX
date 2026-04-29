from __future__ import annotations

from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook


NEW_SOURCE_PATH = Path(r"C:\Users\paul.loja\Downloads\propuesta-categorizacion-productos-base.xlsx")
OLD_PROPOSAL_PATH = Path(r"C:\Users\paul.loja\PYPROYECTOS\dashboard_v2\docs\bodega\propuesta-categorizacion-productos-base.xlsx")
OUTPUT_PATH = Path(r"C:\Users\paul.loja\PYPROYECTOS\dashboard_v2\docs\bodega\propuesta-categorizacion-productos-base.xlsx")

PRODUCT_SHEET = "Productos"


def clean(value: object) -> str:
    if value is None:
        return ""
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text


def normalized_key(value: object) -> str:
    return clean(value).upper()


def slug_code(value: str) -> str:
    text = normalized_key(value)
    text = text.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U").replace("Ñ", "N")
    text = re.sub(r"[^A-Z0-9]+", "_", text).strip("_")
    return text


def is_real_ingredient(value: str) -> bool:
    normalized = normalized_key(value)
    return normalized not in {"", "_", "-", "N/A", "NA", "NO APLICA", "NO APLICA / N-A", "NULL"}


@dataclass
class UnitProposal:
    canonical_code: str
    display_name: str
    dimension: str
    decimal_precision: int
    rationale: str


UNIT_PROPOSALS: dict[str, UnitProposal] = {
    "UN": UnitProposal("UN", "Unidad", "Unidad", 0, "Unidad individual"),
    "U": UnitProposal("UN", "Unidad", "Unidad", 0, "Alias de unidad"),
    "UND": UnitProposal("UN", "Unidad", "Unidad", 0, "Alias de unidad"),
    "CC": UnitProposal("CC", "Centímetro cúbico", "Volumen", 2, "Equivalente funcional a mililitro"),
    "GR": UnitProposal("GR", "Gramo", "Peso", 2, "Unidad de masa"),
    "KG": UnitProposal("KG", "Kilogramo", "Peso", 2, "Unidad de masa"),
    "RO": UnitProposal("RO", "Rollo", "Unidad", 0, "Presentación comercial por rollo"),
    "MT": UnitProposal("MT", "Metro", "Longitud", 2, "Unidad de longitud"),
    "LT": UnitProposal("LT", "Litro", "Volumen", 2, "Unidad de volumen"),
    "PR": UnitProposal("PAR", "Par", "Unidad", 0, "Par de elementos"),
    "PAR": UnitProposal("PAR", "Par", "Unidad", 0, "Par de elementos"),
    "GA": UnitProposal("GL", "Galón", "Volumen", 2, "Galón líquido"),
    "GL": UnitProposal("GL", "Galón", "Volumen", 2, "Galón líquido"),
    "GN": UnitProposal("GL", "Galón", "Volumen", 2, "Inferido por contexto de pinturas y aceites"),
    "SET": UnitProposal("SET", "Set", "Unidad", 0, "Juego o conjunto"),
    "LB": UnitProposal("LB", "Libra", "Peso", 2, "Unidad de masa imperial"),
    "CJ": UnitProposal("CJ", "Caja", "Unidad", 0, "Presentación por caja"),
    "M3": UnitProposal("M3", "Metro cúbico", "Volumen", 3, "Unidad cúbica de volumen"),
    "MT3": UnitProposal("M3", "Metro cúbico", "Volumen", 3, "Alias de metro cúbico"),
    "SACO": UnitProposal("SACO", "Saco", "Unidad", 0, "Presentación por saco"),
    "SA": UnitProposal("SACO", "Saco", "Unidad", 0, "Inferido por contexto de sal y gallinaza"),
    "RM": UnitProposal("RESMA", "Resma", "Unidad", 0, "Presentación de papel por resma"),
    "RESMA": UnitProposal("RESMA", "Resma", "Unidad", 0, "Presentación de papel por resma"),
    "PK": UnitProposal("PK", "Paquete", "Unidad", 0, "Presentación por paquete"),
    "P3": UnitProposal("P3", "Pie cúbico", "Volumen", 3, "Unidad cúbica imperial"),
    "YR": UnitProposal("YR", "Yarda", "Longitud", 2, "Unidad de longitud imperial"),
    "PL": UnitProposal("PL", "Pliego", "Unidad", 0, "Hoja o pliego"),
    "FR": UnitProposal("FR", "Frasco", "Unidad", 0, "Contenedor tipo frasco"),
    "": UnitProposal("UN", "Unidad", "Unidad", 0, "Valor faltante; propuesta conservadora temporal"),
}


FAMILY_NORMALIZATION = {
    "MANTENIMIENTOS": "Mantenimientos",
    "MATERIAL DE EMPAQUE": "Material de Empaque",
    "MATERIAL EMPAQUE": "Material de Empaque",
    "MATERIAL EMP TAPAS": "Material de Empaque",
    "INSUMO EMPAQUE": "Material de Empaque",
    "PESTICIDAS": "Pesticidas",
    "FERTILIZANTES": "Fertilizantes",
    "SEGURIDAD INDUSTRIAL": "Seguridad Industrial",
    "REPUESTOS TRACTOR DE FUMIGACION": "Repuestos Tractor de Fumigacion",
    "INSUMOS DE POSCOSECHA": "Insumos de Poscosecha",
    "PLANTAS COMPRADAS": "Plantas Compradas",
    "PLANTAS COMPRADAS CULTIVOS VARIOS": "Plantas Compradas",
    "SUSTRATOS": "Sustratos",
    "MATERIALES CONSTRUCCION": "Materiales de Construccion",
    "LABORATORIO": "Laboratorio",
    "INSUMOS DE COCINA": "Insumos de Cocina",
    "DESINFECTANTE DE SUELOS": "Desinfectante de Suelos",
    "MATERIALES": "Materiales",
    "VARIOS": "Varios",
}


def canonical_family(raw_family: str) -> str:
    normalized = " ".join(normalized_key(raw_family).split())
    return FAMILY_NORMALIZATION.get(normalized, clean(raw_family) or "Sin familia")


def normalize_old_subfamily(value: str) -> str:
    normalized = canonical_family(value)
    if normalized == "Material de Empaque":
        return normalized
    return clean(value)


OLD_SUBFAMILY_FALLBACKS: dict[tuple[str, str], str] = {
    ("Material de Empaque", "Implementos y repuestos"): "Accesorios de empaque",
    ("Material de Empaque", "Aseo y desinfeccion"): "Accesorios de empaque",
    ("Material de Empaque", "Cafeteria y bienestar"): "Accesorios de empaque",
    ("Material de Empaque", "Capuchon monoorientado"): "Capuchones",
    ("Material de Empaque", "Capuchon microperforado"): "Capuchones",
    ("Material de Empaque", "Etiqueta adhesiva"): "Etiquetas y rotulado",
    ("Material de Empaque", "Etiqueta termica"): "Etiquetas y rotulado",
    ("Pesticidas", "Implementos y repuestos"): "Otros agroquimicos",
    ("Pesticidas", "Reguladores, herbicidas y otros agroquimicos"): "Coadyuvantes y otros agroquimicos",
    ("Fertilizantes", "Implementos y repuestos"): "Fertilizantes solubles y mezclas",
    ("Fertilizantes", "Sales y correctores"): "Micronutrientes y sulfatos",
    ("Fertilizantes", "Bioestimulantes y biologicos"): "Bioestimulantes y biologicos",
    ("Fertilizantes", "Acondicionadores de solucion"): "Acidos y correctores",
    ("Insumos de Poscosecha", "Implementos y repuestos"): "Utensilios y accesorios",
    ("Insumos de Poscosecha", "Aseo y desinfeccion"): "Sanitizacion y desinfeccion",
    ("Insumos de Poscosecha", "Cafeteria y bienestar"): "Utensilios y accesorios",
    ("Insumos de Poscosecha", "Acondicionadores de solucion"): "Preservantes y soluciones florales",
    ("Insumos de Poscosecha", "Tintes base agua"): "Colorantes florales",
    ("Seguridad Industrial", "Guantes y proteccion de manos"): "Guantes y proteccion de manos",
    ("Seguridad Industrial", "Botas y proteccion corporal"): "Proteccion corporal y calzado",
    ("Seguridad Industrial", "Proteccion respiratoria y visual"): "Proteccion respiratoria y visual",
    ("Seguridad Industrial", "Aseo y desinfeccion"): "Higiene y limpieza operativa",
    ("Seguridad Industrial", "Implementos y repuestos"): "Dotacion y seguridad",
    ("Mantenimientos", "Implementos y repuestos"): "Ferreteria y repuestos",
    ("Mantenimientos", "Riego y conexiones"): "Riego y fumigacion",
    ("Mantenimientos", "Guantes y proteccion de manos"): "Consumibles de mantenimiento",
    ("Mantenimientos", "Proteccion respiratoria y visual"): "Consumibles de mantenimiento",
    ("Mantenimientos", "Cafeteria y bienestar"): "Consumibles de mantenimiento",
    ("Mantenimientos", "Acondicionadores de solucion"): "Consumibles de mantenimiento",
    ("Sustratos", "Bioestimulantes y biologicos"): "Complementos organicos",
    ("Laboratorio", "Implementos y repuestos"): "Insumos de laboratorio",
    ("Plantas Compradas", "Implementos y repuestos"): "Material vegetal",
}


def normalized_old_subfamily(family: str, old_subfamily: str) -> str:
    cleaned = clean(old_subfamily)
    if not cleaned:
        return ""
    return OLD_SUBFAMILY_FALLBACKS.get((family, cleaned), cleaned)


def infer_subfamily(family: str, code: str, description: str, old_subfamily: str) -> tuple[str, str, str]:
    desc = normalized_key(description)
    code_prefix = re.match(r"^[A-Z]+", normalized_key(code))
    prefix = code_prefix.group(0) if code_prefix else ""
    old_sub = normalized_old_subfamily(family, old_subfamily)

    def result(name: str, confidence: str, rule: str) -> tuple[str, str, str]:
        return name, confidence, rule

    if family == "Material de Empaque":
        if "CAPUCH" in desc:
            return result("Capuchones", "alta", "keyword capuchon")
        if "TAPA" in desc:
            return result("Tapas y cubiertas", "alta", "keyword tapa")
        if "FONDO" in desc or "CARTON" in desc:
            return result("Fondos y cartones", "alta", "keyword fondo/carton")
        if "ETIQUETA" in desc:
            return result("Etiquetas y rotulado", "alta", "keyword etiqueta")
        if any(term in desc for term in ["FUNDA", "BOLSA", "PLASTIC", "POLIETILENO"]):
            return result("Fundas y plasticos", "alta", "keyword funda/plastico")
        if any(term in desc for term in ["CINTA", "MASKING", "SCOTCH", "ADHES"]):
            return result("Cintas y adhesivos", "alta", "keyword cinta/adhesivo")
        if "LIGA" in desc or "BANDA" in desc:
            return result("Ligas y amarres", "alta", "keyword liga/banda")
        if any(term in desc for term in ["GRAPA", "FLEJE", "SUNCH"]):
            return result("Grapas y sujecion", "alta", "keyword grapa/sujecion")
        if any(term in desc for term in ["PAPEL", "RESMA", "PERIODICO", "SEPARADOR"]):
            return result("Papeles y separadores", "media", "keyword papel")
        if any(term in desc for term in ["PALO", "VARA"]):
            return result("Palos y soportes", "media", "keyword palo/vara")
        if any(term in desc for term in ["CAJA", "BANDEJA"]):
            return result("Cajas y bandejas", "media", "keyword caja/bandeja")
        if old_sub:
            return result(old_sub, "media", "fallback subfamilia historica")
        return result("Accesorios de empaque", "baja", "fallback familia empaque")

    if family == "Pesticidas":
        if prefix == "PF" or any(term in desc for term in ["CAPTAN", "DACONIL", "ANTRACOL", "DITHANE", "TOPAS", "ROVRAL", "VITAVAX", "PHYTON", "KOCIDE", "PREVICUR", "TERRACLOR", "POLYRAM"]):
            return result("Fungicidas", "alta", "prefijo PF o keyword fungicida")
        if prefix == "PI" or any(term in desc for term in ["MESUROL", "RESCATE", "FULMINANTE", "ACETAMIPRID", "CIROMAZINA", "ABAMECTINA", "IMIDACLOPRID", "MOSCA"]):
            return result("Insecticidas y acaricidas", "alta", "prefijo PI o keyword insecticida")
        if "GLIFO" in desc or "HERBI" in desc:
            return result("Herbicidas", "alta", "keyword herbicida")
        if any(term in desc for term in ["GIB", "GA3", "HORM", "REGUL", "AUXINA"]):
            return result("Reguladores y hormonales", "media", "keyword regulador")
        if prefix in {"PH", "PJ", "PN"}:
            return result("Coadyuvantes y otros agroquimicos", "media", f"prefijo {prefix}")
        if old_sub:
            return result(old_sub, "media", "fallback subfamilia historica")
        return result("Otros agroquimicos", "baja", "fallback familia pesticidas")

    if family == "Fertilizantes":
        if any(term in desc for term in ["UREA", "NITRATO DE AMONIO", "SULFATO DE AMONIO"]):
            return result("Nitrogenados", "alta", "keyword nitrogenado")
        if any(term in desc for term in ["FOSFATO", "FOSFORICO", "ROCA FOSFORICA", "MAP", "DAP"]):
            return result("Fosfatados", "alta", "keyword fosfatado")
        if any(term in desc for term in ["POTASIO", "MURIATO", "KCL"]):
            return result("Potasicos", "alta", "keyword potasico")
        if any(term in desc for term in ["CALCIO", "MAGNESIO", "YESO"]):
            return result("Calcio y magnesio", "alta", "keyword calcio/magnesio")
        if any(term in desc for term in ["COBRE", "ZINC", "FERROSO", "HIERRO", "MANGANESO", "BORO", "MICRO"]):
            return result("Micronutrientes y sulfatos", "alta", "keyword micronutriente")
        if any(term in desc for term in ["GALLINAZA", "ABONO", "ORGANICO", "COMPOST", "CHAMPI", "CAL", "CARBONATO"]):
            return result("Organicos y enmiendas", "media", "keyword organico/enmienda")
        if any(term in desc for term in ["ACIDO", "BUFFER", "CORRECTOR"]):
            return result("Acidos y correctores", "media", "keyword acido/corrector")
        if old_sub:
            return result(old_sub, "media", "fallback subfamilia historica")
        return result("Fertilizantes solubles y mezclas", "baja", "fallback familia fertilizantes")

    if family == "Insumos de Poscosecha":
        if any(term in desc for term in ["COLORANTE", "CENTIMO", "TINTE", "HOT PINK", "VIOLETA", "VERDE", "AMARILLO"]):
            return result("Colorantes florales", "alta", "keyword colorante")
        if any(term in desc for term in ["EVERFLOR", "FLOWER FOOD", "LIFE SHINE", "EVERLASTING", "CRUDEX"]):
            return result("Preservantes y soluciones florales", "alta", "keyword preservante floral")
        if any(term in desc for term in ["HIPOCLORITO", "CLORO", "AGUA OXIGENADA", "STARNER"]):
            return result("Sanitizacion y desinfeccion", "alta", "keyword sanitizacion")
        if any(term in desc for term in ["ROJO DE FENOL", "ORTOTOLIDINA", "CLARIFICADOR", "ARQUIMFLOC", "KIT"]):
            return result("Tratamiento y control de agua", "media", "keyword control agua")
        if any(term in desc for term in ["RECIPIENTE", "MEDIDOR", "CAJAS", "BANDEJAS"]):
            return result("Utensilios y accesorios", "media", "keyword utensilio")
        if old_sub:
            return result(old_sub, "media", "fallback subfamilia historica")
        return result("Insumos operativos de poscosecha", "baja", "fallback familia poscosecha")

    if family == "Seguridad Industrial":
        if "GUANTE" in desc:
            return result("Guantes y proteccion de manos", "alta", "keyword guante")
        if any(term in desc for term in ["BOTA", "DELANTAL", "OVEROL", "MANDIL", "CHAQUETA"]):
            return result("Proteccion corporal y calzado", "alta", "keyword proteccion corporal")
        if any(term in desc for term in ["MASCAR", "RESPIR", "CARTUCHO", "LENTE", "GAFA", "CASCO", "ARNES", "OREJERA"]):
            return result("Proteccion respiratoria y visual", "alta", "keyword proteccion respiratoria")
        if any(term in desc for term in ["JABON", "DESINFECT", "PAPEL", "SERVIL", "TOALLA", "ESCOBA", "ALCOHOL", "DETERGENTE", "CEPILLO", "LIMON", "SAL EN GRANO"]):
            return result("Higiene y limpieza operativa", "media", "keyword higiene")
        if old_sub:
            return result(old_sub, "media", "fallback subfamilia historica")
        return result("Dotacion y seguridad", "baja", "fallback familia seguridad")

    if family == "Mantenimientos":
        if any(term in desc for term in ["ACEITE", "GRASA", "DIESEL", "REFRIGERANTE", "R-22", "R-410", "LUBR"]):
            return result("Lubricantes y combustibles", "alta", "keyword lubricante/combustible")
        if any(term in desc for term in ["PINTURA", "ESMALTE", "ANTICORROS", "THINNER"]):
            return result("Pinturas y recubrimientos", "alta", "keyword pintura")
        if any(term in desc for term in ["VALVULA", "VALVULA", "MANGUERA", "PVC", "UNION", "CODO", "TEE", "BOQUILLA", "BOMBA", "ASPERSOR", "CP3", "REGULADOR", "TUBO"]):
            return result("Riego y fumigacion", "alta", "keyword riego/fumigacion")
        if any(term in desc for term in ["CABLE", "LED", "CAPACITOR", "BREAKER", "CONTACTOR", "FOCO", "BOMBILLO", "ENCHUFE"]):
            return result("Electricidad e iluminacion", "alta", "keyword electricidad")
        if any(term in desc for term in ["BALANZA", "MANOMETRO", "MEDIDOR", "TERMOMETRO", "PH", "CONDUCTIVIDAD", "MULTIMETRO"]):
            return result("Medicion y control", "alta", "keyword medicion")
        if any(term in desc for term in ["ALAMBRE", "CLAVO", "TORNILLO", "TUERCA", "RODAMIENTO", "CADENA", "CORREA", "BANDA", "POLEA", "BUJE", "PERNO", "PLATINA", "ANGULO"]):
            return result("Ferreteria y repuestos", "media", "keyword ferreteria")
        if any(term in desc for term in ["PALA", "MACHETE", "AZAD", "RASTRILLO", "TIJERA", "ESCOBADORA", "BARRETA", "COMBO", "PICO"]):
            return result("Herramientas manuales", "alta", "keyword herramienta manual")
        if any(term in desc for term in ["LIJA", "CEPILLO", "SILICON", "PEGA", "AGUA DESTILADA", "LIMPIADOR", "SOLUCION"]):
            return result("Consumibles de mantenimiento", "media", "keyword consumible mantenimiento")
        if old_sub:
            return result(old_sub, "media", "fallback subfamilia historica")
        return result("Mantenimiento general", "baja", "fallback familia mantenimiento")

    if family == "Repuestos Tractor de Fumigacion":
        if any(term in desc for term in ["BOQUILLA", "FILTRO"]):
            return result("Boquillas y filtros", "alta", "keyword boquilla/filtro")
        if any(term in desc for term in ["EMBRAGUE", "PLATO DE PRESION", "PLATO "]):
            return result("Transmision y embrague", "alta", "keyword embrague")
        if any(term in desc for term in ["ENGRANAJE", "EJE"]):
            return result("Engranajes y ejes", "alta", "keyword engranaje/eje")
        if any(term in desc for term in ["FRENO", "PORTADOR", "MONTAJE"]):
            return result("Montajes y frenos", "media", "keyword montaje/freno")
        if any(term in desc for term in ["ACEITE", "SELLA", "SELLO"]):
            return result("Sellos y lubricacion", "media", "keyword sello/lubricacion")
        return result("Repuestos de tractor", "baja", "fallback familia tractor")

    if family == "Sustratos":
        if "BANDEJA" in desc or "CELDA" in desc or "HUECO" in desc:
            return result("Bandejas de propagacion", "alta", "keyword bandeja/celda")
        if any(term in desc for term in ["CASCARILLA", "CASCAJO", "FIBRA DE COCO", "PINO"]):
            return result("Agregados y fibras", "alta", "keyword agregado/fibra")
        if any(term in desc for term in ["SUSTRATO", "KLASMANN", "ECOGREEN", "SEED"]):
            return result("Mezclas de sustrato", "alta", "keyword sustrato")
        if any(term in desc for term in ["ABONO", "COMPOST"]):
            return result("Complementos organicos", "media", "keyword abono/compost")
        return result("Sustratos y propagacion", "baja", "fallback familia sustratos")

    if family == "Plantas Compradas":
        if any(term in desc for term in ["ESQUEJE", "SIN RAIZ", "SEMILLA"]):
            return result("Esquejes y material vegetal", "alta", "keyword esqueje/semilla")
        if any(term in desc for term in ["ZANTEDESCHIA", "ORQUID", "ANTUR", "BROMELIA", "DELPHINIUM"]):
            return result("Plantas ornamentales", "alta", "keyword ornamental")
        if any(term in desc for term in ["PLANTA MADRE", "VARIEDADES", "ARANDANO", "CHAMPI"]):
            return result("Plantas madre y variedades", "media", "keyword planta madre/variedad")
        return result("Material vegetal", "baja", "fallback familia plantas")

    if family == "Laboratorio":
        if any(term in desc for term in ["VASO", "PROBETA", "VIDRIO", "BOECO"]):
            return result("Vidrieria y medicion", "alta", "keyword vidrieria")
        if any(term in desc for term in ["BISTURI", "HOJA"]):
            return result("Material cortopunzante", "alta", "keyword bisturi")
        if any(term in desc for term in ["SOLUCION", "REACTIVO"]):
            return result("Reactivos y soluciones", "media", "keyword reactivo")
        return result("Insumos de laboratorio", "baja", "fallback familia laboratorio")

    if family == "Materiales de Construccion":
        if "CEMENTO" in desc:
            return result("Cementos y aglomerantes", "alta", "keyword cemento")
        if any(term in desc for term in ["ANGULO", "PLATINA", "PERFIL", "TUBO"]):
            return result("Perfiles y metalmecanica", "alta", "keyword perfil/metalmecanica")
        return result("Materiales de construccion general", "baja", "fallback familia construccion")

    if family == "Insumos de Cocina":
        return result("Consumo y cafeteria", "media", "familia unica")

    if family == "Desinfectante de Suelos":
        return result("Desinfectantes de suelo", "media", "familia unica")

    if family == "Materiales":
        return result("Materiales varios", "baja", "familia residual")

    if family == "Varios":
        return result("Productos en prueba", "baja", "familia residual")

    if old_sub:
        return result(old_sub, "media", "fallback subfamilia historica")

    return result("Por clasificar", "baja", "sin regla")


def read_sheet_rows(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, object]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    data: list[dict[str, object]] = []
    for row in rows:
        if not any(value not in (None, "") for value in row):
            continue
        record = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
        data.append(record)
    wb.close()
    return headers, data


def copy_sheet(source_wb, target_wb, sheet_name: str) -> None:
    if sheet_name not in source_wb.sheetnames:
        return
    source_ws = source_wb[sheet_name]
    target_ws = target_wb.create_sheet(sheet_name)
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell._style = copy(cell._style)
            if cell.number_format:
                new_cell.number_format = cell.number_format
            if cell.alignment:
                new_cell.alignment = copy(cell.alignment)
            if cell.font:
                new_cell.font = copy(cell.font)
            if cell.fill:
                new_cell.fill = copy(cell.fill)
            if cell.border:
                new_cell.border = copy(cell.border)
    for key, value in source_ws.column_dimensions.items():
        target_ws.column_dimensions[key].width = value.width
    for key, value in source_ws.row_dimensions.items():
        target_ws.row_dimensions[key].height = value.height


def load_old_index() -> dict[str, dict[str, object]]:
    _, rows = read_sheet_rows(OLD_PROPOSAL_PATH, PRODUCT_SHEET)
    result: dict[str, dict[str, object]] = {}
    for row in rows:
        code = clean(row.get("codigo"))
        if code:
            result[normalized_key(code)] = row
    return result


def make_products_rows() -> list[dict[str, object]]:
    _, rows = read_sheet_rows(NEW_SOURCE_PATH, PRODUCT_SHEET)
    old_index = load_old_index()
    products: list[dict[str, object]] = []

    for row in rows:
        code = clean(row.get("codigo"))
        if not code:
            continue

        source_family = clean(row.get("familia_propuesta"))
        family = canonical_family(source_family)
        old = old_index.get(normalized_key(code), {})
        ingredient = clean(old.get("ingrd_activo"))
        if not is_real_ingredient(ingredient):
            ingredient = ""

        raw_unit = clean(row.get("unidad"))
        unit = UNIT_PROPOSALS.get(normalized_key(raw_unit), UNIT_PROPOSALS[""])
        old_subfamily = clean(old.get("subfamilia_propuesta"))
        subfamily, confidence, rule = infer_subfamily(
            family=family,
            code=code,
            description=clean(row.get("descripcion")),
            old_subfamily=old_subfamily,
        )

        requires_revision = (
            "SI"
            if confidence == "baja" or not raw_unit or not clean(row.get("descripcion"))
            else "NO"
        )

        products.append(
            {
                "source_sheet": clean(row.get("source_sheet")) or "Productos",
                "codigo": code,
                "descripcion": clean(row.get("descripcion")),
                "unidad_fuente": raw_unit,
                "unidad_propuesta": unit.canonical_code,
                "unidad_nombre_propuesta": unit.display_name,
                "dimension_unidad_propuesta": unit.dimension,
                "precision_unidad_propuesta": unit.decimal_precision,
                "ingrd_activo": ingredient,
                "active_component_mode_propuesto": "applies" if is_real_ingredient(ingredient) else "na",
                "familia_fuente": source_family,
                "familia_propuesta": family,
                "subfamilia_propuesta": subfamily,
                "rama_propuesta": f"{family} > {subfamily}",
                "confianza": confidence,
                "criterio_asignacion": rule,
                "requiere_revision": requires_revision,
                "activity_id_1": clean(row.get("activity_id_1")),
                "activity_name_1": clean(row.get("activity_name_1")),
                "cost_area_1": clean(row.get("cost_area_1")),
                "sub_cost_center_1": clean(row.get("sub_cost_center_1")),
                "activity_id_2": clean(row.get("activity_id_2")),
                "activity_name_2": clean(row.get("activity_name_2")),
                "cost_area_2": clean(row.get("cost_area_2")),
                "sub_cost_center_2": clean(row.get("sub_cost_center_2")),
                "activity_id_3": clean(row.get("activity_id_3")),
                "activity_name_3": clean(row.get("activity_name_3")),
                "cost_area_3": clean(row.get("cost_area_3")),
                "sub_cost_center_3": clean(row.get("sub_cost_center_3")),
                "nota_asignacion_actividad": clean(row.get("nota_asignacion_actividad")),
                "subfamilia_historica": old_subfamily,
            }
        )

    return products


def write_table(ws, headers: list[str], rows: list[dict[str, object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


def main() -> None:
    products = make_products_rows()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    product_headers = [
        "source_sheet",
        "codigo",
        "descripcion",
        "unidad_fuente",
        "unidad_propuesta",
        "unidad_nombre_propuesta",
        "dimension_unidad_propuesta",
        "precision_unidad_propuesta",
        "ingrd_activo",
        "active_component_mode_propuesto",
        "familia_fuente",
        "familia_propuesta",
        "subfamilia_propuesta",
        "rama_propuesta",
        "confianza",
        "criterio_asignacion",
        "requiere_revision",
        "activity_id_1",
        "activity_name_1",
        "cost_area_1",
        "sub_cost_center_1",
        "activity_id_2",
        "activity_name_2",
        "cost_area_2",
        "sub_cost_center_2",
        "activity_id_3",
        "activity_name_3",
        "cost_area_3",
        "sub_cost_center_3",
        "nota_asignacion_actividad",
        "subfamilia_historica",
    ]
    ws_products = wb.create_sheet("Productos")
    write_table(ws_products, product_headers, products)

    branch_counter = Counter((row["familia_propuesta"], row["subfamilia_propuesta"]) for row in products)
    branch_rows = [
        {
            "familia_propuesta": family,
            "subfamilia_propuesta": subfamily,
            "productos_asignados": count,
            "rama_propuesta": f"{family} > {subfamily}",
        }
        for (family, subfamily), count in sorted(branch_counter.items())
    ]
    ws_branches = wb.create_sheet("Ramas")
    write_table(
        ws_branches,
        ["familia_propuesta", "subfamilia_propuesta", "productos_asignados", "rama_propuesta"],
        branch_rows,
    )

    family_groups: dict[str, list[str]] = defaultdict(list)
    for row in products:
        family_groups[row["familia_propuesta"]].append(row["subfamilia_propuesta"])
    family_rows = []
    for family in sorted(family_groups):
        distinct_sub = sorted(set(family_groups[family]))
        family_rows.append(
            {
                "familia_propuesta": family,
                "productos_asignados": sum(1 for row in products if row["familia_propuesta"] == family),
                "subfamilias_detectadas": ", ".join(distinct_sub),
            }
        )
    ws_families = wb.create_sheet("Familias")
    write_table(ws_families, ["familia_propuesta", "productos_asignados", "subfamilias_detectadas"], family_rows)

    unit_source_counter: Counter[str] = Counter(normalized_key(row["unidad_fuente"]) for row in products)
    unit_rows = []
    for source_code, count in sorted(unit_source_counter.items()):
        proposal = UNIT_PROPOSALS.get(source_code, UNIT_PROPOSALS[""])
        unit_rows.append(
            {
                "unidad_fuente": source_code,
                "unidad_propuesta": proposal.canonical_code,
                "unidad_nombre_propuesta": proposal.display_name,
                "dimension_unidad_propuesta": proposal.dimension,
                "precision_unidad_propuesta": proposal.decimal_precision,
                "productos_asignados": count,
                "criterio": proposal.rationale,
            }
        )
    ws_units = wb.create_sheet("Unidades")
    write_table(
        ws_units,
        [
            "unidad_fuente",
            "unidad_propuesta",
            "unidad_nombre_propuesta",
            "dimension_unidad_propuesta",
            "precision_unidad_propuesta",
            "productos_asignados",
            "criterio",
        ],
        unit_rows,
    )

    source_wb = load_workbook(NEW_SOURCE_PATH, read_only=False, data_only=False)
    for sheet_name in ("ActividadesFuente", "CentrosCostos"):
        copy_sheet(source_wb, wb, sheet_name)
    source_wb.close()

    summary_rows = [
        {"metrica": "productos_propuestos", "valor": len(products)},
        {"metrica": "familias_unicas", "valor": len(set(row["familia_propuesta"] for row in products))},
        {"metrica": "subfamilias_unicas", "valor": len(set(row["subfamilia_propuesta"] for row in products))},
        {"metrica": "requiere_revision_si", "valor": sum(1 for row in products if row["requiere_revision"] == "SI")},
        {"metrica": "ingredientes_recuperados", "valor": sum(1 for row in products if is_real_ingredient(str(row["ingrd_activo"])))},
    ]
    ws_summary = wb.create_sheet("Resumen")
    write_table(ws_summary, ["metrica", "valor"], summary_rows)

    wb.save(OUTPUT_PATH)
    print(f"Workbook generado en: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
