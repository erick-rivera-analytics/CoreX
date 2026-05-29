from __future__ import annotations

import csv
import json
import re
from datetime import datetime
from difflib import get_close_matches
from pathlib import Path

import psycopg
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(r"C:\Users\paul.loja\AppData\Local\Temp\CoreX_bodega_validate")
ENV_PATH = ROOT / ".env.local"
DATA_PATH = ROOT / "outputs" / "fumigacion_homologacion_20260525" / "fumigacion_homologacion_data.json"
OUTPUT_DIR = ROOT / "outputs" / "fumigacion_homologacion_20260525"
CSV_PATH = OUTPUT_DIR / "fumigacion_bodega_altas_sugeridas.csv"
JSON_PATH = OUTPUT_DIR / "fumigacion_bodega_altas_sugeridas.json"
XLSX_PATH = OUTPUT_DIR / "fumigacion_bodega_altas_sugeridas.xlsx"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).upper()


def load_env() -> dict[str, str]:
    values: dict[str, str] = {}
    for raw in ENV_PATH.read_text(encoding="utf-8").splitlines():
        raw = raw.strip()
        if not raw or raw.startswith("#") or "=" not in raw:
            continue
        key, value = raw.split("=", 1)
        values[key] = value
    return values


def load_bodega_reference() -> tuple[dict[str, dict[str, str]], list[str]]:
    env = load_env()
    conn = psycopg.connect(
        host=env["DATABASE_HOST"],
        port=int(env["DATABASE_PORT"]),
        dbname=env.get("BODEGA_DATABASE_NAME", "db_storageroom"),
        user=env["DATABASE_USER"],
        password=env["DATABASE_PASSWORD"],
    )
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                select
                  upper(trim(p.product_name)) as product_name,
                  upper(trim(p.product_code)) as product_code,
                  upper(trim(coalesce(u.unit_code, ''))) as unit_code,
                  trim(coalesce(c.category_path_label, '')) as category_path_label
                from (
                  select
                    p.product_id,
                    p.product_code,
                    p.product_name,
                    p.base_unit_id,
                    p.category_id
                  from public.sr_dim_product_profile_scd2 p
                  where p.is_current = true
                    and p.is_valid = true
                ) p
                left join (
                  select
                    unit_id,
                    unit_code
                  from public.sr_dim_unit_profile_scd2
                  where is_current = true
                    and is_valid = true
                ) u on u.unit_id = p.base_unit_id
                left join (
                  with recursive category_tree as (
                    select
                      c.category_id,
                      c.category_name,
                      c.parent_category_id,
                      c.category_name::text as path_label
                    from public.sr_dim_category_profile_scd2 c
                    where c.is_current = true
                      and c.is_valid = true
                      and c.parent_category_id is null
                    union all
                    select
                      child.category_id,
                      child.category_name,
                      child.parent_category_id,
                      (parent.path_label || ' / ' || child.category_name)::text as path_label
                    from public.sr_dim_category_profile_scd2 child
                    inner join category_tree parent on parent.category_id = child.parent_category_id
                    where child.is_current = true
                      and child.is_valid = true
                  )
                  select category_id, path_label as category_path_label
                  from category_tree
                ) c on c.category_id = p.category_id
                order by upper(trim(p.product_name)), upper(trim(p.product_code))
                """
            )
            rows = cur.fetchall()
    finally:
        conn.close()

    by_name: dict[str, dict[str, str]] = {}
    names: list[str] = []
    for product_name, product_code, unit_code, category_path_label in rows:
        by_name[product_name] = {
            "product_name": product_name,
            "product_code": product_code,
            "unit_code": unit_code,
            "category_path_label": category_path_label,
        }
        names.append(product_name)
    return by_name, names


def build_rows() -> list[dict[str, object]]:
    payload = json.loads(DATA_PATH.read_text(encoding="utf-8"))
    faltantes = payload.get("faltantes_bodega", [])
    by_name, all_names = load_bodega_reference()

    rows: list[dict[str, object]] = []
    for record in faltantes:
        product_name = normalize_text(record.get("producto_excel"))
        suggestions = get_close_matches(product_name, all_names, n=3, cutoff=0.75)
        suggestion_rows = [by_name[name] for name in suggestions if name in by_name]

        rows.append(
            {
                "producto_excel": product_name,
                "familias_objetivo_requeridas": ", ".join(record.get("familias_objetivo_requeridas", [])),
                "actividades_objetivo_requeridas": ", ".join(record.get("actividades_objetivo_requeridas", [])),
                "programas_detectados": ", ".join(record.get("programas_detectados", [])),
                "semanas_iso_ejemplo": ", ".join(str(value) for value in record.get("semanas_iso_ejemplo", [])),
                "cantidades_ejemplo": ", ".join(str(value) for value in record.get("cantidades_ejemplo", [])),
                "sugerencia_1_nombre_bodega": suggestion_rows[0]["product_name"] if len(suggestion_rows) > 0 else "",
                "sugerencia_1_codigo_bodega": suggestion_rows[0]["product_code"] if len(suggestion_rows) > 0 else "",
                "sugerencia_1_unidad": suggestion_rows[0]["unit_code"] if len(suggestion_rows) > 0 else "",
                "sugerencia_1_categoria": suggestion_rows[0]["category_path_label"] if len(suggestion_rows) > 0 else "",
                "sugerencia_2_nombre_bodega": suggestion_rows[1]["product_name"] if len(suggestion_rows) > 1 else "",
                "sugerencia_2_codigo_bodega": suggestion_rows[1]["product_code"] if len(suggestion_rows) > 1 else "",
                "sugerencia_2_unidad": suggestion_rows[1]["unit_code"] if len(suggestion_rows) > 1 else "",
                "sugerencia_2_categoria": suggestion_rows[1]["category_path_label"] if len(suggestion_rows) > 1 else "",
                "nuevo_product_code_propuesto": "",
                "nuevo_product_name_propuesto": product_name,
                "unidad_base_confirmada": "",
                "categoria_bodega_confirmada": "",
                "componente_activo_mode": "na",
                "componente_activo_nombre": "",
                "descripcion_propuesta": "Alta sugerida desde Programa de Fumigacion CoreX.",
                "is_active": "SI",
                "decision_usuario": "",
                "observaciones": "",
            }
        )

    rows.sort(key=lambda item: str(item["producto_excel"]))
    return rows


def write_csv(rows: list[dict[str, object]]) -> None:
    if not rows:
        return
    with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_json(rows: list[dict[str, object]]) -> None:
    JSON_PATH.write_text(
        json.dumps(
            {
                "generated_at": datetime.now().isoformat(),
                "total_rows": len(rows),
                "rows": rows,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def write_xlsx(rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Altas sugeridas"

    headers = list(rows[0].keys()) if rows else []
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    fill = PatternFill(fill_type="solid", fgColor="1D4ED8")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font

    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        ws.column_dimensions[column_cells[0].column_letter].width = min(width, 42)

    wb.save(XLSX_PATH)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = build_rows()
    write_csv(rows)
    write_json(rows)
    write_xlsx(rows)
    print(json.dumps({
        "total_rows": len(rows),
        "csv": str(CSV_PATH),
        "json": str(JSON_PATH),
        "xlsx": str(XLSX_PATH),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
