from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from datetime import datetime
from difflib import get_close_matches
from pathlib import Path

import psycopg
from openpyxl import load_workbook


ROOT = Path(r"C:\Users\paul.loja\AppData\Local\Temp\CoreX_bodega_validate")
SOURCE_PATH = ROOT / "docs" / "fumigacion_program_source_2608.xlsx"
OUTPUT_DIR = ROOT / "outputs" / "fumigacion_homologacion_20260525"
ENV_PATH = ROOT / ".env.local"

SOURCE_SHEETS = [
    ("NORMAL", "Fumigacion revisado 2339 rev 46", "LANZA_NORMAL"),
    ("DRON", "Fumigacion revisado 2339 Dron", "DRON"),
    ("LANZAS", "Fumigacion2508-Lanzasupereficie", "LANZAS_EFICIENTES"),
]

NOISE_TOKENS = {"#REF!", "SI"}

FUMIGATION_FAMILIES = {
    "LANZA_NORMAL": ["FMGYP", "FMGYPA1", "FMGYPA2", "FMGYPM1", "FMGYPM2", "FMGYPAC2"],
    "LANZAS_EFICIENTES": ["FMGYPEF"],
    "DRON": ["03VAFIFMG", "03VAFIFMGL"],
}

FUMIGATION_ACTIVITY_SET = {
    activity
    for activities in FUMIGATION_FAMILIES.values()
    for activity in activities
}


def normalize_text(value: object) -> str:
    if not isinstance(value, str):
        return ""
    return re.sub(r"\s+", " ", value).strip().upper()


def load_env() -> dict[str, str]:
    values: dict[str, str] = {}
    for raw in ENV_PATH.read_text(encoding="utf-8").splitlines():
        raw = raw.strip()
        if not raw or raw.startswith("#") or "=" not in raw:
            continue
        key, value = raw.split("=", 1)
        values[key] = value
    return values


def load_bodega_catalog() -> tuple[dict[str, list[tuple[str, str]]], set[str]]:
    env = load_env()
    conn = psycopg.connect(
        host=env["DATABASE_HOST"],
        port=int(env["DATABASE_PORT"]),
        dbname=env.get("BODEGA_DATABASE_NAME", "db_storageroom"),
        user=env["DATABASE_USER"],
        password=env["DATABASE_PASSWORD"],
    )
    try:
        cur = conn.cursor()
        cur.execute(
            """
            select
              upper(trim(p.product_code)) as product_code,
              upper(trim(p.product_name)) as product_name,
              upper(trim(u.activity_id)) as activity_id
            from public.sr_dim_product_profile_scd2 p
            left join public.sr_bridge_product_usage_scd2 u
              on u.product_id = p.product_id
             and u.is_current = true
             and u.is_valid = true
            where p.is_current = true
              and p.is_valid = true
            """,
        )
        rows = cur.fetchall()
    finally:
        conn.close()

    by_name: dict[str, list[tuple[str, str]]] = defaultdict(list)
    all_names: set[str] = set()
    for product_code, product_name, activity_id in rows:
        by_name[product_name].append((product_code, activity_id))
        all_names.add(product_name)
    return by_name, all_names


def detect_program_type(header: str) -> str:
    header_upper = header.upper()
    if "MADRES" in header_upper:
        return "CAMAS_MADRES"
    if "ENDURECIMIENTO" in header_upper:
        return "ENDURECIMIENTO"
    return "FUMIGACION_GENERAL"


def extract_sheet_records(
    workbook,
    sheet_code: str,
    sheet_name: str,
    target_family: str,
    bodega_by_name: dict[str, list[tuple[str, str]]],
    all_names: set[str],
) -> list[dict[str, object]]:
    ws = workbook[sheet_name]
    starts = [
        c
        for c in range(1, ws.max_column + 1)
        if isinstance(ws.cell(5, c).value, int)
        and 1000 <= ws.cell(5, c).value <= 9999
        and c >= 16
    ]

    blocks: list[tuple[int, int, str, str]] = []
    for index, start in enumerate(starts):
        end = starts[index + 1] - 1 if index + 1 < len(starts) else ws.max_column
        header = ""
        subheader = ""
        for probe in range(start, min(end, start + 5) + 1):
            candidate = ws.cell(1, probe).value
            if isinstance(candidate, str) and candidate.strip():
                header = re.sub(r"\s+", " ", candidate).strip()
                break
        for probe in range(start, min(end, start + 5) + 1):
            candidate = ws.cell(2, probe).value
            if isinstance(candidate, str) and candidate.strip():
                subheader = re.sub(r"\s+", " ", candidate).strip()
                break
        blocks.append((start, end, header, subheader))

    records: list[dict[str, object]] = []
    seen: set[tuple[str, str, str]] = set()
    for start, end, header, subheader in blocks:
        program_type = detect_program_type(header)
        if program_type != "FUMIGACION_GENERAL":
            continue
        for row_index in range(5, ws.max_row + 1):
            week = ws.cell(row_index, start).value
            if not isinstance(week, int):
                continue
            for column_index in range(start + 1, end + 1, 2):
                product_name = normalize_text(ws.cell(row_index, column_index).value)
                quantity_value = ws.cell(row_index, column_index + 1).value if column_index + 1 <= end else None
                if not product_name or product_name in NOISE_TOKENS:
                    continue
                key = (product_name, header, subheader)
                if key in seen:
                    continue
                seen.add(key)

                matches = bodega_by_name.get(product_name, [])
                exact_exists = bool(matches)
                current_activities = sorted({activity for _, activity in matches if activity})
                required_activities = FUMIGATION_FAMILIES[target_family]
                missing_activities = [activity for activity in required_activities if activity not in current_activities]
                has_target_activity = not missing_activities
                has_any_fumigation_activity = any(activity in FUMIGATION_ACTIVITY_SET for activity in current_activities)
                suggested_names = get_close_matches(product_name, sorted(all_names), n=3, cutoff=0.82) if not exact_exists else []

                if not exact_exists:
                    review_category = "CREAR_EN_BODEGA"
                elif not has_target_activity:
                    review_category = "ASIGNAR_ACTIVIDAD_FUMIGACION"
                else:
                    review_category = "YA_CUBIERTO"

                records.append(
                    {
                        "fuente_tipo": sheet_code,
                        "fuente_hoja": sheet_name,
                        "categoria_revision": review_category,
                        "tipo_programa": program_type,
                        "programa_header": header,
                        "programa_subheader": subheader,
                        "producto_excel": product_name,
                        "semana_iso_ejemplo": week,
                        "cantidad_ejemplo": quantity_value,
                        "existe_nombre_exacto_bodega": "SI" if exact_exists else "NO",
                        "tiene_actividad_fumigacion": "SI" if has_any_fumigation_activity else "NO",
                        "tiene_actividad_objetivo": "SI" if has_target_activity else "NO",
                        "codigos_bodega_exactos": sorted({code for code, _ in matches}),
                        "actividades_actuales_bodega": current_activities,
                        "sugerencias_nombre_bodega": suggested_names,
                        "familia_objetivo_sugerida": target_family,
                        "actividades_objetivo_requeridas": required_activities,
                        "actividades_faltantes_bodega": missing_activities,
                        "actividad_confirmada": "",
                        "categoria_bodega_propuesta": "",
                        "crear_producto_nuevo": "SI" if not exact_exists else "NO",
                        "requiere_asignacion_actividad": "SI" if exact_exists and not has_target_activity else "NO",
                        "decision_usuario": "",
                        "observaciones": "",
                    },
                )

    records.sort(
        key=lambda item: (
            str(item["categoria_revision"]),
            str(item["tipo_programa"]),
            str(item["producto_excel"]),
            str(item["programa_header"]),
        ),
    )
    return records


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    if not rows:
        return
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def build_consolidated_records(records: list[dict[str, object]]) -> list[dict[str, object]]:
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in records:
        grouped[str(record["producto_excel"])].append(record)

    consolidated: list[dict[str, object]] = []
    for product_name, product_records in grouped.items():
        exact_exists = any(record["existe_nombre_exacto_bodega"] == "SI" for record in product_records)
        current_activities = sorted({
            activity
            for record in product_records
            for activity in record["actividades_actuales_bodega"]
        })
        required_activities = sorted({
            activity
            for record in product_records
            for activity in record["actividades_objetivo_requeridas"]
            if activity
        })
        required_families = sorted({
            str(record["familia_objetivo_sugerida"])
            for record in product_records
            if str(record["familia_objetivo_sugerida"])
        })
        has_all_required_activities = exact_exists and all(
            activity in current_activities for activity in required_activities
        )
        missing_activities = [
            activity for activity in required_activities
            if activity not in current_activities
        ]

        if not exact_exists:
            review_category = "CREAR_EN_BODEGA"
        elif not has_all_required_activities:
            review_category = "ASIGNAR_ACTIVIDAD_FUMIGACION"
        else:
            review_category = "YA_CUBIERTO"

        consolidated.append(
            {
                "categoria_revision": review_category,
                "producto_excel": product_name,
                "fuentes_presentes": sorted({str(record["fuente_tipo"]) for record in product_records}),
                "hojas_fuente": sorted({str(record["fuente_hoja"]) for record in product_records}),
                "tipos_programa": sorted({str(record["tipo_programa"]) for record in product_records}),
                "programas_detectados": sorted({
                    str(record["programa_header"])
                    for record in product_records
                    if str(record["programa_header"])
                }),
                "subprogramas_detectados": sorted({
                    str(record["programa_subheader"])
                    for record in product_records
                    if str(record["programa_subheader"])
                }),
                "semanas_iso_ejemplo": sorted({
                    int(record["semana_iso_ejemplo"])
                    for record in product_records
                    if isinstance(record["semana_iso_ejemplo"], int)
                }),
                "cantidades_ejemplo": sorted({
                    str(record["cantidad_ejemplo"])
                    for record in product_records
                    if record["cantidad_ejemplo"] not in (None, "")
                }),
                "existe_nombre_exacto_bodega": "SI" if exact_exists else "NO",
                "tiene_actividad_fumigacion": "SI" if any(
                    record["tiene_actividad_fumigacion"] == "SI" for record in product_records
                ) else "NO",
                "codigos_bodega_exactos": sorted({
                    code
                    for record in product_records
                    for code in record["codigos_bodega_exactos"]
                }),
                "actividades_actuales_bodega": current_activities,
                "sugerencias_nombre_bodega": sorted({
                    suggestion
                    for record in product_records
                    for suggestion in record["sugerencias_nombre_bodega"]
                }),
                "familias_objetivo_requeridas": required_families,
                "actividades_objetivo_requeridas": required_activities,
                "actividades_faltantes_bodega": missing_activities,
                "actividad_confirmada": "",
                "categoria_bodega_propuesta": "",
                "crear_producto_nuevo": "SI" if not exact_exists else "NO",
                "requiere_asignacion_actividad": "SI" if exact_exists and not has_all_required_activities else "NO",
                "decision_usuario": "",
                "observaciones": "",
            },
        )

    consolidated.sort(
        key=lambda item: (
            str(item["categoria_revision"]),
            str(item["producto_excel"]),
        ),
    )
    return consolidated


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(SOURCE_PATH, data_only=True, read_only=False)
    bodega_by_name, all_names = load_bodega_catalog()

    by_source: dict[str, list[dict[str, object]]] = {}
    all_records: list[dict[str, object]] = []

    for sheet_code, sheet_name, target_family in SOURCE_SHEETS:
        records = extract_sheet_records(workbook, sheet_code, sheet_name, target_family, bodega_by_name, all_names)
        by_source[sheet_code] = records
        all_records.extend(records)

    consolidated_records = build_consolidated_records(all_records)
    faltantes_bodega = [
        record for record in consolidated_records if record["categoria_revision"] == "CREAR_EN_BODEGA"
    ]
    sin_actividad_fumigacion = [
        record for record in consolidated_records if record["categoria_revision"] == "ASIGNAR_ACTIVIDAD_FUMIGACION"
    ]

    payload = {
        "generated_at": datetime.now().isoformat(),
        "source_sheets": [
            {
                "sheet_code": sheet_code,
                "sheet_name": sheet_name,
                "target_family": target_family,
                "target_activities": FUMIGATION_FAMILIES[target_family],
            }
            for sheet_code, sheet_name, target_family in SOURCE_SHEETS
        ],
        "records": all_records,
        "records_consolidated": consolidated_records,
        "by_source": by_source,
        "faltantes_bodega": faltantes_bodega,
        "sin_actividad_fumigacion": sin_actividad_fumigacion,
    }

    (OUTPUT_DIR / "fumigacion_homologacion_data.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    write_csv(OUTPUT_DIR / "universo_completo.csv", all_records)
    write_csv(OUTPUT_DIR / "universo_consolidado.csv", consolidated_records)
    write_csv(OUTPUT_DIR / "faltantes_bodega.csv", faltantes_bodega)
    write_csv(OUTPUT_DIR / "sin_actividad_fumigacion.csv", sin_actividad_fumigacion)
    for source_code, rows in by_source.items():
        write_csv(OUTPUT_DIR / f"{source_code.lower()}_trabajo.csv", rows)

    print(f"OK {OUTPUT_DIR}")
    print(f"TOTAL {len(all_records)}")
    print(f"UNICOS {len(consolidated_records)}")
    print(f"FALTANTES {len(faltantes_bodega)}")
    print(f"SIN_ACTIVIDAD {len(sin_actividad_fumigacion)}")
    for source_code, rows in by_source.items():
        print(f"{source_code} {len(rows)}")


if __name__ == "__main__":
    main()
