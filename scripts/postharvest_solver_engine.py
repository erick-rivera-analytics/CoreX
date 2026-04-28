from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
import itertools
import os
from pathlib import Path
import shutil
from typing import Any

import numpy as np
import pandas as pd
import pulp

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional workbook dependency
    load_workbook = None


WORKBOOK_FILENAME = "Copia de Solver Tallos.xlsm"
MASTER_FILENAME = "data/sku_master.csv"
DATE_COLUMNS = ["fecha_1", "fecha_2", "fecha_3", "fecha_4", "fecha_5"]
SKU_COLUMN = "sku"
MASTER_COLUMNS = [
    SKU_COLUMN,
    "peso_ideal_bunch",
    "tallos_min",
    "tallos_max",
    "peso_min_objetivo",
    "peso_max_objetivo",
    "max_grados_objetivo",
]


@dataclass
class WorkbookDefaults:
    sku_master: pd.DataFrame
    orders: pd.DataFrame
    availability: pd.DataFrame
    settings: dict[str, float]
    workbook_path: Path
    master_path: Path


@dataclass
class Stage1Result:
    orders: pd.DataFrame
    summary: dict[str, float]


@dataclass
class Stage2Result:
    orders: pd.DataFrame
    net_tallos: pd.DataFrame
    gross_tallos: pd.DataFrame
    mallas_raw: pd.DataFrame
    mallas_display: pd.DataFrame
    availability: pd.DataFrame
    summary: dict[str, float]
    solver_meta: dict[str, float | str]


@dataclass
class PipelineResult:
    stage1: Stage1Result
    stage2: Stage2Result


def default_workbook_path() -> Path:
    return Path(__file__).resolve().parent / WORKBOOK_FILENAME


def default_master_path() -> Path:
    return Path(__file__).resolve().parent / MASTER_FILENAME


def excel_round(value: float, digits: int = 0) -> float:
    factor = 10**digits
    scaled = float(value) * factor
    if scaled >= 0:
        rounded = np.floor(scaled + 0.5)
    else:
        rounded = np.ceil(scaled - 0.5)
    return float(rounded / factor)


def excel_round_array(values: np.ndarray | pd.Series, digits: int = 0) -> np.ndarray:
    array = np.asarray(values, dtype=float)
    factor = 10**digits
    scaled = array * factor
    rounded = np.where(scaled >= 0, np.floor(scaled + 0.5), np.ceil(scaled - 0.5))
    return rounded / factor


def numeric(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def normalize_label(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)) and float(value).is_integer():
        return str(int(value))
    return str(value).strip()


def parse_tallos_max(raw_range: Any, tallos_min: int) -> int:
    text = normalize_label(raw_range)
    if "-" not in text:
        return tallos_min
    left, right = [part.strip() for part in text.split("-", 1)]
    try:
        left_value = int(float(left))
        right_value = int(float(right))
    except ValueError:
        return tallos_min
    if left_value == tallos_min and right_value >= tallos_min:
        return right_value
    return tallos_min


def sanitize_master(df: pd.DataFrame) -> pd.DataFrame:
    master = df.copy()
    if SKU_COLUMN not in master.columns and "variedad" in master.columns:
        master = master.rename(columns={"variedad": SKU_COLUMN})

    if "peso_min_objetivo" not in master.columns and "ratio_min_objetivo" in master.columns:
        master["peso_min_objetivo"] = (
            pd.to_numeric(master.get("ratio_min_objetivo"), errors="coerce").fillna(0.97)
            * pd.to_numeric(master.get("peso_ideal_bunch"), errors="coerce").fillna(0.0)
        )
    if "peso_max_objetivo" not in master.columns and "ratio_max_objetivo" in master.columns:
        master["peso_max_objetivo"] = (
            pd.to_numeric(master.get("ratio_max_objetivo"), errors="coerce").fillna(1.03)
            * pd.to_numeric(master.get("peso_ideal_bunch"), errors="coerce").fillna(0.0)
        )
    for column in MASTER_COLUMNS:
        if column not in master.columns:
            master[column] = np.nan

    master[SKU_COLUMN] = master[SKU_COLUMN].map(normalize_label)
    master = master[master[SKU_COLUMN] != ""].copy()
    master["peso_ideal_bunch"] = pd.to_numeric(master["peso_ideal_bunch"], errors="coerce").fillna(0.0)
    master["tallos_min"] = (
        pd.to_numeric(master["tallos_min"], errors="coerce").fillna(0.0).map(lambda value: int(excel_round(value, 0)))
    )
    master["tallos_max"] = (
        pd.to_numeric(master["tallos_max"], errors="coerce").fillna(master["tallos_min"]).map(lambda value: int(excel_round(value, 0)))
    )
    master["tallos_max"] = np.maximum(master["tallos_max"], master["tallos_min"])
    master["peso_min_objetivo"] = pd.to_numeric(master["peso_min_objetivo"], errors="coerce").fillna(
        master["peso_ideal_bunch"] * 0.97
    )
    master["peso_max_objetivo"] = pd.to_numeric(master["peso_max_objetivo"], errors="coerce").fillna(
        master["peso_ideal_bunch"] * 1.03
    )
    master["max_grados_objetivo"] = (
        pd.to_numeric(master["max_grados_objetivo"], errors="coerce").fillna(3.0).map(lambda value: int(excel_round(value, 0)))
    )
    master["peso_min_objetivo"] = np.minimum(master["peso_min_objetivo"], master["peso_ideal_bunch"])
    master["peso_max_objetivo"] = np.maximum(master["peso_max_objetivo"], master["peso_ideal_bunch"])
    master["peso_max_objetivo"] = np.maximum(master["peso_max_objetivo"], master["peso_min_objetivo"])
    master["max_grados_objetivo"] = np.maximum(master["max_grados_objetivo"], 1)
    if master[SKU_COLUMN].duplicated().any():
        duplicates = sorted(master.loc[master[SKU_COLUMN].duplicated(), SKU_COLUMN].unique().tolist())
        raise ValueError(f"Hay SKUs repetidos en la tabla maestra: {', '.join(duplicates)}")
    master = master[MASTER_COLUMNS].sort_values(SKU_COLUMN).reset_index(drop=True)
    return master


def sanitize_orders(df: pd.DataFrame, master_df: pd.DataFrame) -> pd.DataFrame:
    orders = df.copy()
    if SKU_COLUMN not in orders.columns and "variedad" in orders.columns:
        orders = orders.rename(columns={"variedad": SKU_COLUMN})
    if SKU_COLUMN not in orders.columns:
        raise ValueError("La tabla de pedidos debe incluir la columna 'sku'.")
    orders[SKU_COLUMN] = orders[SKU_COLUMN].map(normalize_label)
    for column in DATE_COLUMNS:
        if column not in orders.columns:
            orders[column] = 0.0
        orders[column] = pd.to_numeric(orders[column], errors="coerce").fillna(0.0)
    orders = orders[orders[SKU_COLUMN] != ""].copy()
    if orders.empty:
        raise ValueError("Debes ingresar al menos un pedido.")
    orders = orders.groupby(SKU_COLUMN, as_index=False)[DATE_COLUMNS].sum()
    for column in DATE_COLUMNS:
        orders[column] = orders[column].map(lambda value: int(excel_round(value, 0)))
    orders["pedido_total"] = orders[DATE_COLUMNS].sum(axis=1).map(lambda value: int(excel_round(value, 0)))
    orders = orders[orders["pedido_total"] > 0].copy()
    if orders.empty:
        raise ValueError("Todos los pedidos estan en cero.")
    merged = orders.merge(master_df, on=SKU_COLUMN, how="left", validate="many_to_one")
    missing = merged.loc[merged["peso_ideal_bunch"].isna(), SKU_COLUMN].tolist()
    if missing:
        raise ValueError(
            "Hay pedidos sin SKU en la tabla maestra: " + ", ".join(sorted(set(missing)))
        )
    return merged.reset_index(drop=True)


def sanitize_availability(df: pd.DataFrame, desperdicio: float) -> pd.DataFrame:
    availability = df.copy()
    if "grado" not in availability.columns:
        raise ValueError("La tabla de disponibilidad debe incluir la columna 'grado'.")
    availability["grado"] = (
        pd.to_numeric(availability["grado"], errors="coerce").fillna(0.0).map(lambda value: int(excel_round(value, 0)))
    )
    availability = availability[availability["grado"] > 0].copy()
    for column in DATE_COLUMNS:
        if column not in availability.columns:
            availability[column] = 0.0
        availability[column] = pd.to_numeric(availability[column], errors="coerce").fillna(0.0)
    availability["peso_tallo_seed"] = pd.to_numeric(availability["peso_tallo_seed"], errors="coerce").fillna(0.0)
    availability["mallas_totales"] = availability[DATE_COLUMNS].sum(axis=1)
    availability["tallos_brutos"] = availability["mallas_totales"] * 20.0
    availability["tallos_netos"] = excel_round_array(availability["tallos_brutos"] * (1.0 - desperdicio), 0)
    availability["peso_total_gestionable"] = availability["tallos_netos"] * availability["peso_tallo_seed"]
    availability = availability.sort_values("grado").reset_index(drop=True)
    return availability


def save_master_table(master_df: pd.DataFrame, path: str | Path | None = None) -> Path:
    target = Path(path) if path is not None else default_master_path()
    target.parent.mkdir(parents=True, exist_ok=True)
    sanitized = sanitize_master(master_df)
    sanitized.to_csv(target, index=False, encoding="utf-8-sig")
    return target


def generate_default_master_from_workbook(workbook_path: Path) -> pd.DataFrame:
    if load_workbook is None:
        raise RuntimeError("OpenPyXL no esta disponible para construir defaults desde workbook.")
    wb = load_workbook(workbook_path, data_only=True, keep_vba=True)
    ws = wb["REGISTRO_PEDIDOS"]
    records: list[dict[str, Any]] = []
    for row in range(9, 24):
        sku = normalize_label(ws[f"B{row}"].value or ws[f"O{row}"].value)
        tallos_min = int(excel_round(numeric(ws[f"E{row}"].value), 0))
        tallos_max = parse_tallos_max(ws[f"F{row}"].value, tallos_min)
        peso_ideal = numeric(ws[f"C{row}"].value)
        records.append(
            {
                SKU_COLUMN: sku,
                "peso_ideal_bunch": peso_ideal,
                "tallos_min": tallos_min,
                "tallos_max": tallos_max,
                "peso_min_objetivo": peso_ideal * 0.97,
                "peso_max_objetivo": peso_ideal * 1.03,
                "max_grados_objetivo": 3,
            }
        )
    return sanitize_master(pd.DataFrame(records))


def load_sku_master(master_path: str | Path | None = None, workbook_path: str | Path | None = None) -> tuple[pd.DataFrame, Path]:
    target = Path(master_path) if master_path is not None else default_master_path()
    workbook = Path(workbook_path) if workbook_path is not None else default_workbook_path()
    if target.exists():
        df = pd.read_csv(target)
        return sanitize_master(df), target
    default_master = generate_default_master_from_workbook(workbook)
    save_master_table(default_master, target)
    return default_master, target


def load_workbook_defaults(
    workbook_path: str | Path | None = None,
    master_path: str | Path | None = None,
) -> WorkbookDefaults:
    if load_workbook is None:
        raise RuntimeError("OpenPyXL no esta disponible para cargar defaults desde workbook.")
    workbook = Path(workbook_path) if workbook_path is not None else default_workbook_path()
    wb = load_workbook(workbook, data_only=True, keep_vba=True)
    ws_reg = wb["REGISTRO_PEDIDOS"]

    orders_records: list[dict[str, Any]] = []
    for row in range(9, 24):
        orders_records.append(
            {
                SKU_COLUMN: normalize_label(ws_reg[f"B{row}"].value or ws_reg[f"O{row}"].value),
                "fecha_1": numeric(ws_reg[f"P{row}"].value),
                "fecha_2": numeric(ws_reg[f"Q{row}"].value),
                "fecha_3": numeric(ws_reg[f"R{row}"].value),
                "fecha_4": numeric(ws_reg[f"S{row}"].value),
                "fecha_5": numeric(ws_reg[f"T{row}"].value),
            }
        )

    weight_seed_by_grade: dict[int, float] = {}
    for row in range(30, 43):
        grade = int(excel_round(numeric(ws_reg[f"G{row}"].value), 0))
        weight_seed_by_grade[grade] = numeric(ws_reg[f"K{row}"].value)

    availability_records: list[dict[str, Any]] = []
    for row in range(9, 22):
        grade = int(excel_round(numeric(ws_reg[f"G{row}"].value), 0))
        availability_records.append(
            {
                "grado": grade,
                "fecha_1": numeric(ws_reg[f"H{row}"].value),
                "fecha_2": numeric(ws_reg[f"I{row}"].value),
                "fecha_3": numeric(ws_reg[f"J{row}"].value),
                "fecha_4": numeric(ws_reg[f"K{row}"].value),
                "fecha_5": numeric(ws_reg[f"L{row}"].value),
                "peso_tallo_seed": weight_seed_by_grade.get(grade, float(grade)),
            }
        )

    sku_master, resolved_master_path = load_sku_master(master_path=master_path, workbook_path=workbook)
    settings = {
        "desperdicio": numeric(ws_reg["C27"].value, 0.13),
    }
    return WorkbookDefaults(
        sku_master=sku_master,
        orders=pd.DataFrame(orders_records),
        availability=pd.DataFrame(availability_records),
        settings=settings,
        workbook_path=workbook,
        master_path=resolved_master_path,
    )


STAGE1_TIME_LIMIT_SECONDS = 5
STAGE2_TIME_LIMIT_SECONDS = 5
OBJECTIVE_TOLERANCE = 1e-4
INTEGRAL_TOLERANCE = 1e-5
WEIGHT_TOLERANCE = 1e-2
MACRO_WEIGHT_MIN_RATIO = 0.97
MACRO_WEIGHT_TARGET_RATIO = 1.0


def sku_allows_euro_stem_flex(sku: Any) -> bool:
    label = normalize_label(sku).upper()
    if not label:
        return False
    if "5X5" in label or "EX" in label:
        return False
    return "750X" in label or "1000X" in label


def stem_shortfall_flex_per_bunch(sku: Any) -> float:
    return 2.0 if sku_allows_euro_stem_flex(sku) else 0.0


def stem_overrun_flex_per_bunch(sku: Any) -> float:
    return 0.0


def dynamic_weight_bounds(ideal_bunch_weight: float) -> tuple[float, float]:
    ideal = max(float(ideal_bunch_weight), 0.0)
    lower_abs = max(ideal * 0.03, 8.0)
    upper_abs = max(ideal * 0.03, 5.0)
    return max(ideal - lower_abs, 0.0), ideal + upper_abs


@lru_cache(maxsize=None)
def _integer_compositions(total: int, parts: int) -> tuple[tuple[int, ...], ...]:
    if parts == 1:
        return ((total,),)
    rows: list[tuple[int, ...]] = []
    for value in range(total + 1):
        for tail in _integer_compositions(total - value, parts - 1):
            rows.append((value, *tail))
    return tuple(rows)


@lru_cache(maxsize=None)
def best_recipe_signature(
    sku: str,
    ideal_bunch_weight: float,
    tallos_min: int,
    tallos_max: int,
    max_grados_objetivo: int,
    grade_weight_items: tuple[tuple[int, float], ...],
) -> tuple[float, float]:
    lower_weight, upper_weight = dynamic_weight_bounds(ideal_bunch_weight)
    min_stems = max(int(tallos_min - stem_shortfall_flex_per_bunch(sku)), 1)
    max_stems = max(int(tallos_max + stem_overrun_flex_per_bunch(sku)), min_stems)
    max_positive_grades = max(int(max_grados_objetivo), 1) + 1
    weights = tuple(float(weight) for _, weight in grade_weight_items)
    best_choice: tuple[float, float, int] | None = None

    for stems_total in range(min_stems, max_stems + 1):
        for grade_count in range(1, min(max_positive_grades, len(weights)) + 1):
            for subset_indexes in itertools.combinations(range(len(weights)), grade_count):
                subset_weights = tuple(weights[index] for index in subset_indexes)
                for counts in _integer_compositions(stems_total, grade_count):
                    if not any(counts):
                        continue
                    weight_total = float(sum(count * weight for count, weight in zip(counts, subset_weights)))
                    range_penalty = max(lower_weight - weight_total, 0.0) + max(weight_total - upper_weight, 0.0)
                    deviation_abs = abs(weight_total - float(ideal_bunch_weight))
                    choice = (range_penalty, deviation_abs, stems_total)
                    if best_choice is None or choice < best_choice:
                        best_choice = choice

    if best_choice is None:
        return float(tallos_min), float(ideal_bunch_weight)
    return float(best_choice[2]), float(ideal_bunch_weight)


def objective_fix_tolerance(opt_value: float) -> float:
    return max(OBJECTIVE_TOLERANCE, abs(float(opt_value)) * 1e-6)


def is_truthy(value: str | None) -> bool:
    if value is None:
        return False
    return value.strip().lower() in {"1", "true", "yes", "on"}


def build_cbc_solver(time_limit: float | None = None) -> pulp.LpSolver:
    solver_path = os.environ.get("POSTHARVEST_CBC_PATH", "").strip()
    if not solver_path:
        solver_path = shutil.which("cbc") or ""

    msg_enabled = is_truthy(os.environ.get("POSTHARVEST_SOLVER_MSG"))

    if solver_path:
        return pulp.COIN_CMD(path=solver_path, msg=msg_enabled, timeLimit=time_limit)

    return pulp.PULP_CBC_CMD(msg=msg_enabled, timeLimit=time_limit)


def status_name(status_code: int) -> str:
    return pulp.LpStatus.get(status_code, str(status_code))


def solve_or_raise(
    problem: pulp.LpProblem,
    solver: pulp.LpSolver,
    error_prefix: str,
    integer_vars: list[pulp.LpVariable] | None = None,
    allow_feasible_incumbent: bool = False,
) -> str:
    status_code = problem.solve(solver)
    resolved_status = status_name(status_code)
    if resolved_status == "Optimal":
        return resolved_status
    if allow_feasible_incumbent and resolved_status == "Not Solved":
        tracked_vars = integer_vars or []
        if tracked_vars:
            values = [pulp.value(variable) for variable in tracked_vars]
            if values and all(
                value is not None and abs(float(value) - round(float(value))) <= INTEGRAL_TOLERANCE
                for value in values
            ):
                return resolved_status
        elif pulp.value(problem.objective) is not None:
            return resolved_status
    raise ValueError(f"{error_prefix}: {resolved_status}")


def solve_pipeline(
    orders_df: pd.DataFrame,
    availability_df: pd.DataFrame,
    settings: dict[str, float],
    master_df: pd.DataFrame,
) -> PipelineResult:
    desperdicio = numeric(settings.get("desperdicio"), 0.0)
    master = sanitize_master(master_df)
    orders = sanitize_orders(orders_df, master)
    availability = sanitize_availability(availability_df, desperdicio)

    n_orders = len(orders)
    n_grades = len(availability)
    active_grade_positions = [
        grade_position
        for grade_position in range(n_grades)
        if float(availability.loc[grade_position, "tallos_netos"]) > 0.5
    ]

    stage1_solver = build_cbc_solver(STAGE1_TIME_LIMIT_SECONDS)
    stage1_problem = pulp.LpProblem("solver_poscosecha_stage1", pulp.LpMaximize)
    stage1_b = {
        order_position: pulp.LpVariable(
            f"stage1_b_{order_position}",
            lowBound=0,
            upBound=int(orders.loc[order_position, "pedido_total"]),
            cat="Integer",
        )
        for order_position in range(n_orders)
    }
    stage1_x = {
        (order_position, grade_position): pulp.LpVariable(
            f"stage1_x_{order_position}_{grade_position}",
            lowBound=0,
            upBound=float(availability.loc[grade_position, "tallos_netos"]),
        )
        for order_position in range(n_orders)
        for grade_position in active_grade_positions
    }
    stage1_stems = {
        order_position: pulp.LpVariable(f"stage1_stems_{order_position}", lowBound=0)
        for order_position in range(n_orders)
    }
    stage1_stem_shortfall = {
        order_position: pulp.LpVariable(f"stage1_stem_shortfall_{order_position}", lowBound=0)
        for order_position in range(n_orders)
    }
    stage1_stem_overrun = {
        order_position: pulp.LpVariable(f"stage1_stem_overrun_{order_position}", lowBound=0)
        for order_position in range(n_orders)
    }
    stage1_weights = {
        order_position: pulp.LpVariable(f"stage1_weight_{order_position}", lowBound=0)
        for order_position in range(n_orders)
    }
    stage1_recipe_stem_gap = {
        order_position: pulp.LpVariable(f"stage1_recipe_stem_gap_{order_position}", lowBound=0)
        for order_position in range(n_orders)
    }
    stage1_macro_floor_shortfall = pulp.LpVariable("stage1_macro_floor_shortfall", lowBound=0)
    stage1_macro_target_low = pulp.LpVariable("stage1_macro_target_low", lowBound=0)
    stage1_macro_target_high = pulp.LpVariable("stage1_macro_target_high", lowBound=0)
    stage1_over_ideal = {
        order_position: pulp.LpVariable(f"stage1_over_ideal_{order_position}", lowBound=0)
        for order_position in range(n_orders)
    }
    stage1_under_ideal = {
        order_position: pulp.LpVariable(f"stage1_under_ideal_{order_position}", lowBound=0)
        for order_position in range(n_orders)
    }
    stage1_fulfilled_by_date = {
        (order_position, date_position): pulp.LpVariable(
            f"stage1_fulfilled_{order_position}_{date_position}",
            lowBound=0,
            upBound=int(orders.loc[order_position, DATE_COLUMNS[date_position]]),
            cat="Integer",
        )
        for order_position in range(n_orders)
        for date_position in range(len(DATE_COLUMNS))
    }

    for order_position in range(n_orders):
        order_shortfall_flex = stem_shortfall_flex_per_bunch(orders.loc[order_position, SKU_COLUMN])
        order_overrun_flex = stem_overrun_flex_per_bunch(orders.loc[order_position, SKU_COLUMN])
        preferred_recipe_stems, _ = best_recipe_signature(
            normalize_label(orders.loc[order_position, SKU_COLUMN]),
            float(orders.loc[order_position, "peso_ideal_bunch"]),
            int(orders.loc[order_position, "tallos_min"]),
            int(orders.loc[order_position, "tallos_max"]),
            int(orders.loc[order_position, "max_grados_objetivo"]),
            tuple(
                (
                    int(availability.loc[grade_position, "grado"]),
                    float(availability.loc[grade_position, "peso_tallo_seed"]),
                )
                for grade_position in active_grade_positions
            ),
        )
        stage1_problem += (
            pulp.lpSum(
                stage1_fulfilled_by_date[order_position, date_position]
                for date_position in range(len(DATE_COLUMNS))
            )
            == stage1_b[order_position]
        ), f"stage1_fulfilled_date_balance_{order_position}"
        stage1_problem += stage1_stems[order_position] == pulp.lpSum(
            stage1_x[order_position, grade_position] for grade_position in active_grade_positions
        ), f"stage1_stems_balance_{order_position}"
        stage1_problem += stage1_weights[order_position] == pulp.lpSum(
            stage1_x[order_position, grade_position] * float(availability.loc[grade_position, "peso_tallo_seed"])
            for grade_position in active_grade_positions
        ), f"stage1_weight_balance_{order_position}"
        stage1_problem += (
            stage1_stems[order_position] + stage1_stem_shortfall[order_position]
            >= float(orders.loc[order_position, "tallos_min"]) * stage1_b[order_position]
        ), f"stage1_min_stems_soft_{order_position}"
        stage1_problem += (
            stage1_stems[order_position] - stage1_stem_overrun[order_position]
            <= float(orders.loc[order_position, "tallos_max"]) * stage1_b[order_position]
        ), f"stage1_max_stems_soft_{order_position}"
        stage1_problem += (
            stage1_stem_shortfall[order_position]
            <= order_shortfall_flex * stage1_b[order_position]
        ), f"stage1_min_stems_soft_cap_{order_position}"
        stage1_problem += (
            stage1_stem_overrun[order_position]
            <= order_overrun_flex * stage1_b[order_position]
        ), f"stage1_max_stems_soft_cap_{order_position}"
        stage1_problem += (
            stage1_recipe_stem_gap[order_position]
            >= stage1_stems[order_position] - preferred_recipe_stems * stage1_b[order_position]
        ), f"stage1_recipe_gap_high_{order_position}"
        stage1_problem += (
            stage1_recipe_stem_gap[order_position]
            >= preferred_recipe_stems * stage1_b[order_position] - stage1_stems[order_position]
        ), f"stage1_recipe_gap_low_{order_position}"
        stage1_problem += (
            stage1_over_ideal[order_position]
            >= stage1_weights[order_position]
            - float(orders.loc[order_position, "peso_ideal_bunch"]) * stage1_b[order_position]
        ), f"stage1_over_ideal_balance_{order_position}"
        stage1_problem += (
            stage1_under_ideal[order_position]
            >= float(orders.loc[order_position, "peso_ideal_bunch"]) * stage1_b[order_position]
            - stage1_weights[order_position]
        ), f"stage1_under_ideal_balance_{order_position}"

    for grade_position in active_grade_positions:
        stage1_problem += (
            pulp.lpSum(stage1_x[order_position, grade_position] for order_position in range(n_orders))
            <= float(availability.loc[grade_position, "tallos_netos"])
        ), f"stage1_capacity_grade_{grade_position}"

    stage1_priority_statuses: dict[str, str] = {}
    stage1_priority_optima: dict[str, float] = {}
    stage1_status_macro = "No aplica"
    stage1_status_recipe = "No aplica"
    stage1_status_overweight = "No aplica"
    stage1_status_ideal = "No aplica"
    stage1_status_stems = "No aplica"
    stage1_macro_violation_opt = 0.0
    stage1_macro_target_deviation_opt = 0.0
    stage1_recipe_gap_opt = 0.0
    stage1_overweight_opt = 0.0
    stage1_ideal_deviation_opt = 0.0
    stage1_stem_violation_opt = 0.0
    for date_position, date_column in enumerate(DATE_COLUMNS):
        priority_expr = pulp.lpSum(
            stage1_fulfilled_by_date[order_position, date_position]
            for order_position in range(n_orders)
        )
        stage1_problem.setObjective(priority_expr)
        stage1_priority_statuses[date_column] = solve_or_raise(
            stage1_problem,
            stage1_solver,
            f"No se pudo priorizar {date_column}",
            integer_vars=list(stage1_b.values()) + list(stage1_fulfilled_by_date.values()),
            allow_feasible_incumbent=True,
        )
        stage1_priority_optima[date_column] = float(pulp.value(priority_expr) or 0.0)
        stage1_problem += (
            priority_expr >= stage1_priority_optima[date_column] - objective_fix_tolerance(stage1_priority_optima[date_column])
        ), f"stage1_fix_priority_{date_column}"

    total_stage1_weight_expr = pulp.lpSum(stage1_weights[order_position] for order_position in range(n_orders))
    fulfilled_ideal_expr = pulp.lpSum(
        stage1_b[order_position] * float(orders.loc[order_position, "peso_ideal_bunch"])
        for order_position in range(n_orders)
    )
    stage1_problem += (
        stage1_macro_floor_shortfall >= MACRO_WEIGHT_MIN_RATIO * fulfilled_ideal_expr - total_stage1_weight_expr
    ), "stage1_macro_floor_balance"
    stage1_problem += (
        stage1_macro_target_low >= MACRO_WEIGHT_TARGET_RATIO * fulfilled_ideal_expr - total_stage1_weight_expr
    ), "stage1_macro_target_low_balance"
    stage1_problem += (
        stage1_macro_target_high >= total_stage1_weight_expr - MACRO_WEIGHT_TARGET_RATIO * fulfilled_ideal_expr
    ), "stage1_macro_target_high_balance"
    stage1_macro_violation_expr = stage1_macro_floor_shortfall
    stage1_macro_target_deviation_expr = stage1_macro_target_low + stage1_macro_target_high
    stage1_problem.setObjective(-stage1_macro_violation_expr)
    stage1_status_macro = solve_or_raise(
        stage1_problem,
        stage1_solver,
        "No se pudo minimizar la desviacion macro de peso por prioridad",
        integer_vars=list(stage1_b.values()) + list(stage1_fulfilled_by_date.values()),
        allow_feasible_incumbent=True,
    )
    stage1_macro_violation_opt = float(pulp.value(stage1_macro_violation_expr) or 0.0)
    stage1_problem += (
        stage1_macro_violation_expr <= stage1_macro_violation_opt + objective_fix_tolerance(stage1_macro_violation_opt)
    ), "stage1_fix_macro_violation"

    stage1_problem.setObjective(-stage1_macro_target_deviation_expr)
    stage1_status_macro = solve_or_raise(
        stage1_problem,
        stage1_solver,
        "No se pudo acercar el peso macro estimado al ideal",
        integer_vars=list(stage1_b.values()) + list(stage1_fulfilled_by_date.values()),
        allow_feasible_incumbent=True,
    )
    stage1_macro_target_deviation_opt = float(pulp.value(stage1_macro_target_deviation_expr) or 0.0)
    stage1_problem += (
        stage1_macro_target_deviation_expr
        <= stage1_macro_target_deviation_opt + objective_fix_tolerance(stage1_macro_target_deviation_opt)
    ), "stage1_fix_macro_target_deviation"

    stage1_recipe_gap_expr = pulp.lpSum(stage1_recipe_stem_gap[order_position] for order_position in range(n_orders))
    stage1_problem.setObjective(-stage1_recipe_gap_expr)
    stage1_status_recipe = solve_or_raise(
        stage1_problem,
        stage1_solver,
        "No se pudo minimizar la desviacion temprana de receta por prioridad",
        integer_vars=list(stage1_b.values()) + list(stage1_fulfilled_by_date.values()),
        allow_feasible_incumbent=True,
    )
    stage1_recipe_gap_opt = float(pulp.value(stage1_recipe_gap_expr) or 0.0)
    stage1_problem += (
        stage1_recipe_gap_expr <= stage1_recipe_gap_opt + objective_fix_tolerance(stage1_recipe_gap_opt)
    ), "stage1_fix_recipe_gap"

    stage1_overweight_expr = pulp.lpSum(stage1_over_ideal[order_position] for order_position in range(n_orders))
    stage1_problem.setObjective(-stage1_overweight_expr)
    stage1_status_overweight = solve_or_raise(
        stage1_problem,
        stage1_solver,
        "No se pudo minimizar el sobrepeso estimado por prioridad",
        integer_vars=list(stage1_b.values()) + list(stage1_fulfilled_by_date.values()),
        allow_feasible_incumbent=True,
    )
    stage1_overweight_opt = float(pulp.value(stage1_overweight_expr) or 0.0)
    stage1_problem += (
        stage1_overweight_expr <= stage1_overweight_opt + objective_fix_tolerance(stage1_overweight_opt)
    ), "stage1_fix_overweight"

    stage1_ideal_deviation_expr = pulp.lpSum(
        stage1_over_ideal[order_position] + stage1_under_ideal[order_position]
        for order_position in range(n_orders)
    )
    stage1_problem.setObjective(-stage1_ideal_deviation_expr)
    stage1_status_ideal = solve_or_raise(
        stage1_problem,
        stage1_solver,
        "No se pudo minimizar la desviacion estimada por pedido respecto al ideal",
        integer_vars=list(stage1_b.values()) + list(stage1_fulfilled_by_date.values()),
        allow_feasible_incumbent=True,
    )
    stage1_ideal_deviation_opt = float(pulp.value(stage1_ideal_deviation_expr) or 0.0)
    stage1_problem += (
        stage1_ideal_deviation_expr <= stage1_ideal_deviation_opt + objective_fix_tolerance(stage1_ideal_deviation_opt)
    ), "stage1_fix_ideal_deviation"

    stage1_stem_violation_expr = pulp.lpSum(
        stage1_stem_shortfall[order_position] + stage1_stem_overrun[order_position]
        for order_position in range(n_orders)
    )
    stage1_problem.setObjective(-stage1_stem_violation_expr)
    stage1_status_stems = solve_or_raise(
        stage1_problem,
        stage1_solver,
        "No se pudo minimizar la desviacion de tallos por prioridad",
        integer_vars=list(stage1_b.values()) + list(stage1_fulfilled_by_date.values()),
        allow_feasible_incumbent=True,
    )
    stage1_stem_violation_opt = float(pulp.value(stage1_stem_violation_expr) or 0.0)
    stage1_problem += (
        stage1_stem_violation_expr <= stage1_stem_violation_opt + objective_fix_tolerance(stage1_stem_violation_opt)
    ), "stage1_fix_stem_violation"

    stage1_problem.setObjective(fulfilled_ideal_expr)
    stage1_status = solve_or_raise(
        stage1_problem,
        stage1_solver,
        "No se pudo resolver la etapa interna de pedidos",
        integer_vars=list(stage1_b.values()),
        allow_feasible_incumbent=True,
    )
    fulfilled_ideal_opt = float(pulp.value(fulfilled_ideal_expr) or 0.0)
    fulfilled_bunches = np.array(
        [int(excel_round(pulp.value(stage1_b[order_position]) or 0.0, 0)) for order_position in range(n_orders)],
        dtype=int,
    )
    fulfilled_by_date_matrix = np.array(
        [
            [
                int(excel_round(pulp.value(stage1_fulfilled_by_date[order_position, date_position]) or 0.0, 0))
                for date_position in range(len(DATE_COLUMNS))
            ]
            for order_position in range(n_orders)
        ],
        dtype=int,
    )

    active_order_positions = [
        order_position for order_position in range(n_orders) if int(fulfilled_bunches[order_position]) > 0
    ]
    stage2_status_pref = "No aplica"
    stage2_status_macro = "No aplica"
    stage2_status_balance = "No aplica"
    stage2_status_recipe = "No aplica"
    stage2_status_overweight = "No aplica"
    stage2_status_ideal = "No aplica"
    stage2_status_stems = "No aplica"
    stage2_status_extra = "No aplica"
    stage2_status_total = "No aplica"
    stage2_status_weight = "No aplica"
    preferred_violation_opt = 0.0
    macro_violation_opt = 0.0
    macro_target_deviation_opt = 0.0
    balance_overweight_opt = 0.0
    recipe_gap_opt = 0.0
    overweight_opt = 0.0
    ideal_deviation_opt = 0.0
    stem_violation_opt = 0.0
    extra_grades_opt = 0.0
    total_grades_opt = 0.0

    net_matrix = np.zeros((n_orders, n_grades), dtype=float)
    weights_used = np.zeros(n_orders, dtype=float)
    slack_low_values = np.zeros(n_orders, dtype=float)
    slack_high_values = np.zeros(n_orders, dtype=float)
    extra_grades_values = np.zeros(n_orders, dtype=float)
    stem_shortfall_values = np.zeros(n_orders, dtype=float)
    stem_overrun_values = np.zeros(n_orders, dtype=float)
    recipe_gap_values = np.zeros(n_orders, dtype=float)

    if active_order_positions:
        stage2_solver = build_cbc_solver(STAGE2_TIME_LIMIT_SECONDS)
        stage2_problem = pulp.LpProblem("solver_poscosecha_stage2", pulp.LpMinimize)
        stage2_x = {
            (order_position, grade_position): pulp.LpVariable(
                f"stage2_x_{order_position}_{grade_position}",
                lowBound=0,
                upBound=float(availability.loc[grade_position, "tallos_netos"]),
            )
            for order_position in active_order_positions
            for grade_position in active_grade_positions
        }
        stage2_u = {
            (order_position, grade_position): pulp.LpVariable(
                f"stage2_u_{order_position}_{grade_position}",
                lowBound=0,
                upBound=1,
                cat="Binary",
            )
            for order_position in active_order_positions
            for grade_position in active_grade_positions
        }
        stage2_stems = {
            order_position: pulp.LpVariable(f"stage2_stems_{order_position}", lowBound=0)
            for order_position in active_order_positions
        }
        stage2_stem_shortfall = {
            order_position: pulp.LpVariable(f"stage2_stem_shortfall_{order_position}", lowBound=0)
            for order_position in active_order_positions
        }
        stage2_stem_overrun = {
            order_position: pulp.LpVariable(f"stage2_stem_overrun_{order_position}", lowBound=0)
            for order_position in active_order_positions
        }
        stage2_recipe_stem_gap = {
            order_position: pulp.LpVariable(f"stage2_recipe_stem_gap_{order_position}", lowBound=0)
            for order_position in active_order_positions
        }
        stage2_weights = {
            order_position: pulp.LpVariable(f"stage2_weight_{order_position}", lowBound=0)
            for order_position in active_order_positions
        }
        stage2_slack_low = {
            order_position: pulp.LpVariable(f"stage2_slack_low_{order_position}", lowBound=0)
            for order_position in active_order_positions
        }
        stage2_slack_high = {
            order_position: pulp.LpVariable(f"stage2_slack_high_{order_position}", lowBound=0)
            for order_position in active_order_positions
        }
        stage2_over_ideal = {
            order_position: pulp.LpVariable(f"stage2_over_ideal_{order_position}", lowBound=0)
            for order_position in active_order_positions
        }
        stage2_under_ideal = {
            order_position: pulp.LpVariable(f"stage2_under_ideal_{order_position}", lowBound=0)
            for order_position in active_order_positions
        }
        stage2_extra_grades = {
            order_position: pulp.LpVariable(f"stage2_extra_grades_{order_position}", lowBound=0)
            for order_position in active_order_positions
        }
        stage2_macro_floor_shortfall = pulp.LpVariable("stage2_macro_floor_shortfall", lowBound=0)
        stage2_macro_target_low = pulp.LpVariable("stage2_macro_target_low", lowBound=0)
        stage2_macro_target_high = pulp.LpVariable("stage2_macro_target_high", lowBound=0)
        stage2_max_deviation_ratio = pulp.LpVariable("stage2_max_deviation_ratio", lowBound=0)

        for order_position in active_order_positions:
            fixed_bunches = int(fulfilled_bunches[order_position])
            order_shortfall_flex = stem_shortfall_flex_per_bunch(orders.loc[order_position, SKU_COLUMN])
            order_overrun_flex = stem_overrun_flex_per_bunch(orders.loc[order_position, SKU_COLUMN])
            dynamic_min_weight, dynamic_max_weight = dynamic_weight_bounds(
                float(orders.loc[order_position, "peso_ideal_bunch"])
            )
            preferred_recipe_stems, _ = best_recipe_signature(
                normalize_label(orders.loc[order_position, SKU_COLUMN]),
                float(orders.loc[order_position, "peso_ideal_bunch"]),
                int(orders.loc[order_position, "tallos_min"]),
                int(orders.loc[order_position, "tallos_max"]),
                int(orders.loc[order_position, "max_grados_objetivo"]),
                tuple(
                    (
                        int(availability.loc[grade_position, "grado"]),
                        float(availability.loc[grade_position, "peso_tallo_seed"]),
                    )
                    for grade_position in active_grade_positions
                ),
            )
            stage2_problem += stage2_stems[order_position] == pulp.lpSum(
                stage2_x[order_position, grade_position] for grade_position in active_grade_positions
            ), f"stage2_stems_balance_{order_position}"
            stage2_problem += stage2_weights[order_position] == pulp.lpSum(
                stage2_x[order_position, grade_position] * float(availability.loc[grade_position, "peso_tallo_seed"])
                for grade_position in active_grade_positions
            ), f"stage2_weight_balance_{order_position}"
            stage2_problem += (
                stage2_stems[order_position] + stage2_stem_shortfall[order_position]
                >= float(orders.loc[order_position, "tallos_min"]) * fixed_bunches
            ), f"stage2_min_stems_soft_{order_position}"
            stage2_problem += (
                stage2_stems[order_position] - stage2_stem_overrun[order_position]
                <= float(orders.loc[order_position, "tallos_max"]) * fixed_bunches
            ), f"stage2_max_stems_soft_{order_position}"
            stage2_problem += (
                stage2_stem_shortfall[order_position]
                <= order_shortfall_flex * fixed_bunches
            ), f"stage2_min_stems_soft_cap_{order_position}"
            stage2_problem += (
                stage2_stem_overrun[order_position]
                <= order_overrun_flex * fixed_bunches
            ), f"stage2_max_stems_soft_cap_{order_position}"
            stage2_problem += (
                stage2_recipe_stem_gap[order_position]
                >= stage2_stems[order_position] - preferred_recipe_stems * fixed_bunches
            ), f"stage2_recipe_gap_high_{order_position}"
            stage2_problem += (
                stage2_recipe_stem_gap[order_position]
                >= preferred_recipe_stems * fixed_bunches - stage2_stems[order_position]
            ), f"stage2_recipe_gap_low_{order_position}"
            stage2_problem += (
                stage2_slack_low[order_position]
                >= dynamic_min_weight * fixed_bunches - stage2_weights[order_position]
            ), f"stage2_soft_low_{order_position}"
            stage2_problem += (
                stage2_slack_high[order_position]
                >= stage2_weights[order_position]
                - dynamic_max_weight * fixed_bunches
            ), f"stage2_soft_high_{order_position}"
            stage2_problem += (
                stage2_over_ideal[order_position]
                >= stage2_weights[order_position]
                - float(orders.loc[order_position, "peso_ideal_bunch"]) * fixed_bunches
            ), f"stage2_over_ideal_{order_position}"
            stage2_problem += (
                stage2_under_ideal[order_position]
                >= float(orders.loc[order_position, "peso_ideal_bunch"]) * fixed_bunches
                - stage2_weights[order_position]
            ), f"stage2_under_ideal_{order_position}"
            stage2_problem += (
                stage2_extra_grades[order_position]
                >= pulp.lpSum(stage2_u[order_position, grade_position] for grade_position in active_grade_positions)
                - float(orders.loc[order_position, "max_grados_objetivo"])
            ), f"stage2_extra_grade_penalty_{order_position}"
            ideal_total = float(orders.loc[order_position, "peso_ideal_bunch"]) * fixed_bunches
            if ideal_total > 0:
                stage2_problem += (
                    stage2_over_ideal[order_position]
                    <= stage2_max_deviation_ratio * ideal_total
                ), f"stage2_max_over_ratio_{order_position}"
                stage2_problem += (
                    stage2_under_ideal[order_position]
                    <= stage2_max_deviation_ratio * ideal_total
                ), f"stage2_max_under_ratio_{order_position}"
            for grade_position in active_grade_positions:
                stage2_problem += (
                    stage2_x[order_position, grade_position]
                    <= float(availability.loc[grade_position, "tallos_netos"])
                    * stage2_u[order_position, grade_position]
                ), f"stage2_link_{order_position}_{grade_position}"

        for grade_position in active_grade_positions:
            stage2_problem += (
                pulp.lpSum(stage2_x[order_position, grade_position] for order_position in active_order_positions)
                <= float(availability.loc[grade_position, "tallos_netos"])
            ), f"stage2_capacity_grade_{grade_position}"

        preferred_violation_expr = pulp.lpSum(
            stage2_slack_low[order_position] + stage2_slack_high[order_position]
            for order_position in active_order_positions
        )
        overweight_expr = pulp.lpSum(
            stage2_over_ideal[order_position]
            for order_position in active_order_positions
        )
        ideal_deviation_expr = pulp.lpSum(
            stage2_over_ideal[order_position] + stage2_under_ideal[order_position]
            for order_position in active_order_positions
        )
        stem_violation_expr = pulp.lpSum(
            stage2_stem_shortfall[order_position] + stage2_stem_overrun[order_position]
            for order_position in active_order_positions
        )
        recipe_gap_expr = pulp.lpSum(
            stage2_recipe_stem_gap[order_position] for order_position in active_order_positions
        )
        extra_grades_expr = pulp.lpSum(stage2_extra_grades[order_position] for order_position in active_order_positions)
        total_grades_expr = pulp.lpSum(
            stage2_u[order_position, grade_position]
            for order_position in active_order_positions
            for grade_position in active_grade_positions
        )
        actual_weight_expr = pulp.lpSum(stage2_weights[order_position] for order_position in active_order_positions)
        resolved_ideal_total = float(
            sum(
                fulfilled_bunches[order_position] * float(orders.loc[order_position, "peso_ideal_bunch"])
                for order_position in active_order_positions
            )
        )
        stage2_problem += (
            stage2_macro_floor_shortfall >= MACRO_WEIGHT_MIN_RATIO * resolved_ideal_total - actual_weight_expr
        ), "stage2_macro_floor_balance"
        stage2_problem += (
            stage2_macro_target_low >= MACRO_WEIGHT_TARGET_RATIO * resolved_ideal_total - actual_weight_expr
        ), "stage2_macro_target_low_balance"
        stage2_problem += (
            stage2_macro_target_high >= actual_weight_expr - MACRO_WEIGHT_TARGET_RATIO * resolved_ideal_total
        ), "stage2_macro_target_high_balance"
        macro_violation_expr = stage2_macro_floor_shortfall
        macro_target_deviation_expr = stage2_macro_target_low + stage2_macro_target_high
        balance_deviation_expr = stage2_max_deviation_ratio

        stage2_problem.setObjective(macro_violation_expr)
        stage2_status_macro = solve_or_raise(
            stage2_problem,
            stage2_solver,
            "No se pudo minimizar la desviacion macro de peso",
            integer_vars=list(stage2_u.values()),
            allow_feasible_incumbent=True,
        )
        macro_violation_opt = float(pulp.value(macro_violation_expr) or 0.0)
        stage2_problem += (
            macro_violation_expr <= macro_violation_opt + objective_fix_tolerance(macro_violation_opt)
        ), "stage2_fix_macro_violation"

        stage2_problem.setObjective(macro_target_deviation_expr)
        stage2_status_weight = solve_or_raise(
            stage2_problem,
            stage2_solver,
            "No se pudo acercar el peso macro final al ideal",
            integer_vars=list(stage2_u.values()),
            allow_feasible_incumbent=True,
        )
        macro_target_deviation_opt = float(pulp.value(macro_target_deviation_expr) or 0.0)
        stage2_problem += (
            macro_target_deviation_expr
            <= macro_target_deviation_opt + objective_fix_tolerance(macro_target_deviation_opt)
        ), "stage2_fix_macro_target_deviation"

        stage2_problem.setObjective(balance_deviation_expr)
        stage2_status_balance = solve_or_raise(
            stage2_problem,
            stage2_solver,
            "No se pudo balancear la desviacion maxima por SKU",
            integer_vars=list(stage2_u.values()),
            allow_feasible_incumbent=True,
        )
        balance_overweight_opt = float(pulp.value(balance_deviation_expr) or 0.0)
        stage2_problem += (
            balance_deviation_expr <= balance_overweight_opt + objective_fix_tolerance(balance_overweight_opt)
        ), "stage2_fix_balance_deviation"

        stage2_problem.setObjective(ideal_deviation_expr)
        stage2_status_ideal = solve_or_raise(
            stage2_problem,
            stage2_solver,
            "No se pudo minimizar la desviacion total respecto al ideal",
            integer_vars=list(stage2_u.values()),
            allow_feasible_incumbent=True,
        )
        ideal_deviation_opt = float(pulp.value(ideal_deviation_expr) or 0.0)
        stage2_problem += (
            ideal_deviation_expr <= ideal_deviation_opt + objective_fix_tolerance(ideal_deviation_opt)
        ), "stage2_fix_ideal_deviation"

        stage2_problem.setObjective(overweight_expr)
        stage2_status_overweight = solve_or_raise(
            stage2_problem,
            stage2_solver,
            "No se pudo minimizar el sobrepeso real residual respecto al ideal",
            integer_vars=list(stage2_u.values()),
            allow_feasible_incumbent=True,
        )
        overweight_opt = float(pulp.value(overweight_expr) or 0.0)
        stage2_problem += (
            overweight_expr <= overweight_opt + objective_fix_tolerance(overweight_opt)
        ), "stage2_fix_overweight"

        stage2_problem.setObjective(recipe_gap_expr)
        stage2_status_recipe = solve_or_raise(
            stage2_problem,
            stage2_solver,
            "No se pudo minimizar la desviacion temprana de receta por pedido",
            integer_vars=list(stage2_u.values()),
            allow_feasible_incumbent=True,
        )
        recipe_gap_opt = float(pulp.value(recipe_gap_expr) or 0.0)
        stage2_problem += (
            recipe_gap_expr <= recipe_gap_opt + objective_fix_tolerance(recipe_gap_opt)
        ), "stage2_fix_recipe_gap"

        stage2_problem.setObjective(preferred_violation_expr)
        stage2_status_pref = solve_or_raise(
            stage2_problem,
            stage2_solver,
            "No se pudo minimizar la desviacion objetivo por pedido",
            integer_vars=list(stage2_u.values()),
            allow_feasible_incumbent=True,
        )
        preferred_violation_opt = float(pulp.value(preferred_violation_expr) or 0.0)
        stage2_problem += (
            preferred_violation_expr <= preferred_violation_opt + objective_fix_tolerance(preferred_violation_opt)
        ), "stage2_fix_preferred_violation"

        stage2_problem.setObjective(stem_violation_expr)
        stage2_status_stems = solve_or_raise(
            stage2_problem,
            stage2_solver,
            "No se pudo minimizar la desviacion de tallos por pedido",
            integer_vars=list(stage2_u.values()),
            allow_feasible_incumbent=True,
        )
        stem_violation_opt = float(pulp.value(stem_violation_expr) or 0.0)
        stage2_problem += (
            stem_violation_expr <= stem_violation_opt + objective_fix_tolerance(stem_violation_opt)
        ), "stage2_fix_stem_violation"

        stage2_problem.setObjective(extra_grades_expr)
        stage2_status_extra = solve_or_raise(
            stage2_problem,
            stage2_solver,
            "No se pudo minimizar el exceso de grados objetivo",
            integer_vars=list(stage2_u.values()),
            allow_feasible_incumbent=True,
        )
        extra_grades_opt = float(pulp.value(extra_grades_expr) or 0.0)
        stage2_problem += (
            extra_grades_expr <= extra_grades_opt + objective_fix_tolerance(extra_grades_opt)
        ), "stage2_fix_extra_grades"

        stage2_problem.setObjective(total_grades_expr)
        stage2_status_total = solve_or_raise(
            stage2_problem,
            stage2_solver,
            "No se pudo minimizar el numero total de grados usados",
            integer_vars=list(stage2_u.values()),
            allow_feasible_incumbent=True,
        )
        total_grades_opt = float(pulp.value(total_grades_expr) or 0.0)

        for order_position in active_order_positions:
            weights_used[order_position] = float(pulp.value(stage2_weights[order_position]) or 0.0)
            slack_low_values[order_position] = float(pulp.value(stage2_slack_low[order_position]) or 0.0)
            slack_high_values[order_position] = float(pulp.value(stage2_slack_high[order_position]) or 0.0)
            extra_grades_values[order_position] = float(pulp.value(stage2_extra_grades[order_position]) or 0.0)
            stem_shortfall_values[order_position] = float(pulp.value(stage2_stem_shortfall[order_position]) or 0.0)
            stem_overrun_values[order_position] = float(pulp.value(stage2_stem_overrun[order_position]) or 0.0)
            recipe_gap_values[order_position] = float(pulp.value(stage2_recipe_stem_gap[order_position]) or 0.0)
            for grade_position in active_grade_positions:
                net_matrix[order_position, grade_position] = float(
                    pulp.value(stage2_x[order_position, grade_position]) or 0.0
                )

    net_matrix = np.maximum(net_matrix, 0.0)
    gross_matrix = np.where(1.0 - desperdicio > 0, net_matrix / (1.0 - desperdicio), 0.0)
    mallas_matrix = gross_matrix / 20.0

    order_names = orders[SKU_COLUMN].tolist()
    grade_labels = availability["grado"].astype(int).tolist()
    stems_used = net_matrix.sum(axis=1)
    requested_bunches = orders["pedido_total"].to_numpy(dtype=int)
    requested_ideal_weight = requested_bunches * orders["peso_ideal_bunch"].to_numpy(dtype=float)
    fulfilled_ideal_weight = fulfilled_bunches * orders["peso_ideal_bunch"].to_numpy(dtype=float)
    requested_min_stems = requested_bunches * orders["tallos_min"].to_numpy(dtype=float)
    adjustments = requested_bunches - fulfilled_bunches
    compliance = np.divide(fulfilled_bunches, requested_bunches, out=np.ones_like(fulfilled_bunches, dtype=float), where=requested_bunches > 0)
    grades_used = (net_matrix > 1e-8).sum(axis=1).astype(int)
    avg_stems_per_bunch = np.divide(stems_used, fulfilled_bunches, out=np.zeros_like(stems_used), where=fulfilled_bunches > 0)
    avg_weight_per_bunch = np.divide(weights_used, fulfilled_bunches, out=np.zeros_like(weights_used), where=fulfilled_bunches > 0)
    avg_stem_weight_real = np.divide(weights_used, stems_used, out=np.zeros_like(weights_used), where=stems_used > 0)
    avg_stem_weight_target = np.divide(
        orders["peso_ideal_bunch"].to_numpy(dtype=float),
        orders["tallos_min"].to_numpy(dtype=float),
        out=np.zeros(n_orders, dtype=float),
        where=orders["tallos_min"].to_numpy(dtype=float) > 0,
    )
    dynamic_bounds = np.array(
        [dynamic_weight_bounds(weight) for weight in orders["peso_ideal_bunch"].to_numpy(dtype=float)],
        dtype=float,
    )
    peso_min_obj = dynamic_bounds[:, 0]
    peso_max_obj = dynamic_bounds[:, 1]
    ratio_real_vs_ideal = np.divide(
        avg_weight_per_bunch,
        orders["peso_ideal_bunch"].to_numpy(dtype=float),
        out=np.ones_like(avg_weight_per_bunch),
        where=(orders["peso_ideal_bunch"].to_numpy(dtype=float) > 0) & (fulfilled_bunches > 0),
    )
    over_under_weight = weights_used - fulfilled_ideal_weight
    over_under_weight_per_bunch = np.where(
        fulfilled_bunches > 0,
        avg_weight_per_bunch - orders["peso_ideal_bunch"].to_numpy(dtype=float),
        0.0,
    )
    over_under_pct = np.where(fulfilled_bunches > 0, ratio_real_vs_ideal - 1.0, 0.0)
    within_objective = (slack_low_values < OBJECTIVE_TOLERANCE) & (slack_high_values < OBJECTIVE_TOLERANCE)
    below_objective = (fulfilled_bunches > 0) & (avg_weight_per_bunch < peso_min_obj - WEIGHT_TOLERANCE)
    above_objective = (fulfilled_bunches > 0) & (avg_weight_per_bunch > peso_max_obj + WEIGHT_TOLERANCE)
    estado_peso = np.full(n_orders, "Fuera de objetivo", dtype=object)
    estado_peso[fulfilled_bunches == 0] = "Sin resolver"
    estado_peso[below_objective] = "Debajo de objetivo"
    estado_peso[above_objective] = "Sobre objetivo"
    estado_peso[(fulfilled_bunches > 0) & within_objective] = "Dentro de objetivo"

    order_summary = orders.copy()
    order_summary[SKU_COLUMN] = order_summary[SKU_COLUMN]
    order_summary["pedido_resuelto"] = fulfilled_bunches
    order_summary["ajuste_bunches"] = adjustments
    order_summary["cumplimiento_bunches"] = compliance
    order_summary["peso_ideal_pedido"] = requested_ideal_weight
    order_summary["peso_ideal_resuelto"] = fulfilled_ideal_weight
    order_summary["peso_min_objetivo"] = peso_min_obj
    order_summary["peso_max_objetivo"] = peso_max_obj
    order_summary["tallos_asignados_netos"] = stems_used
    order_summary["tallos_asignados_brutos"] = gross_matrix.sum(axis=1)
    order_summary["mallas_totales"] = mallas_matrix.sum(axis=1)
    order_summary["tallos_promedio_ramo"] = avg_stems_per_bunch
    order_summary["peso_promedio_tallo_objetivo"] = avg_stem_weight_target
    order_summary["peso_promedio_tallo_real"] = avg_stem_weight_real
    order_summary["peso_real_total"] = weights_used
    order_summary["peso_real_bunch"] = avg_weight_per_bunch
    order_summary["ratio_real_vs_ideal"] = ratio_real_vs_ideal
    order_summary["sobrepeso_pct"] = over_under_pct
    order_summary["estado_peso"] = estado_peso
    order_summary["sobrepeso_bunch"] = over_under_weight_per_bunch
    order_summary["sobrepeso_total"] = over_under_weight
    order_summary["grados_usados"] = grades_used
    order_summary["exceso_grados_objetivo"] = extra_grades_values
    order_summary["desviacion_tallos_bajo"] = stem_shortfall_values
    order_summary["desviacion_tallos_sobre"] = stem_overrun_values
    order_summary["desviacion_tallos_receta"] = recipe_gap_values
    order_summary["desviacion_bajo_objetivo"] = slack_low_values
    order_summary["desviacion_sobre_objetivo"] = slack_high_values
    order_summary["peso_min_dinamico"] = peso_min_obj
    order_summary["peso_max_dinamico"] = peso_max_obj
    order_summary["dentro_rango_objetivo"] = within_objective

    net_tallos = pd.DataFrame(net_matrix, index=order_names, columns=grade_labels)
    gross_tallos = pd.DataFrame(gross_matrix, index=order_names, columns=grade_labels)
    mallas_raw = pd.DataFrame(mallas_matrix, index=order_names, columns=grade_labels)
    mallas_display = pd.DataFrame(
        excel_round_array(mallas_raw.to_numpy(dtype=float), 0).astype(int),
        index=order_names,
        columns=grade_labels,
    )

    availability_summary = availability.copy()
    availability_summary["tallos_usados_netos"] = net_matrix.sum(axis=0)
    availability_summary["tallos_usados_brutos"] = gross_matrix.sum(axis=0)
    availability_summary["mallas_usadas"] = mallas_matrix.sum(axis=0)
    availability_summary["tallos_restantes_netos"] = availability_summary["tallos_netos"] - availability_summary["tallos_usados_netos"]
    availability_summary["peso_usado"] = availability_summary["tallos_usados_netos"] * availability_summary["peso_tallo_seed"]
    availability_summary["peso_restante"] = availability_summary["peso_total_gestionable"] - availability_summary["peso_usado"]

    peso_disponible_total = float(availability_summary["peso_total_gestionable"].sum())
    peso_real_total = float(weights_used.sum())
    peso_ideal_total = float(requested_ideal_weight.sum())
    peso_ideal_resuelto_total = float(fulfilled_ideal_weight.sum())
    tallos_disponibles_netos = float(availability_summary["tallos_netos"].sum())
    tallos_usados_netos = float(net_matrix.sum())
    tallos_objetivo_resueltos = float((fulfilled_bunches * orders["tallos_min"].to_numpy(dtype=float)).sum())
    peso_promedio_tallo_objetivo = (
        float(peso_ideal_resuelto_total / tallos_objetivo_resueltos) if tallos_objetivo_resueltos > 0 else 0.0
    )
    peso_promedio_tallo_disponible = (
        float(peso_disponible_total / tallos_disponibles_netos) if tallos_disponibles_netos > 0 else 0.0
    )
    peso_promedio_tallo_real = float(peso_real_total / tallos_usados_netos) if tallos_usados_netos > 0 else 0.0
    cumplimiento_peso_macro = (
        float(peso_real_total / peso_ideal_resuelto_total) if peso_ideal_resuelto_total > 0 else 1.0
    )

    stage1_summary = {
        "pedido_bunches_total": float(requested_bunches.sum()),
        "pedido_bunches_resuelto": float(fulfilled_bunches.sum()),
        "ajuste_bunches_total": float(adjustments.sum()),
        "peso_ideal_pedido_total": peso_ideal_total,
        "peso_ideal_resuelto_total": peso_ideal_resuelto_total,
        "tallos_pedidos_minimos_total": float(requested_min_stems.sum()),
        "tallos_disponibles_netos": tallos_disponibles_netos,
        "tallos_usados_netos": tallos_usados_netos,
        "tallos_restantes_netos": float(tallos_disponibles_netos - tallos_usados_netos),
    }
    for date_position, date_column in enumerate(DATE_COLUMNS):
        pedido_fecha = float(orders[date_column].sum())
        resuelto_fecha = float(fulfilled_by_date_matrix[:, date_position].sum())
        stage1_summary[f"pedido_{date_column}"] = pedido_fecha
        stage1_summary[f"resuelto_{date_column}"] = resuelto_fecha
        stage1_summary[f"no_realizado_{date_column}"] = float(pedido_fecha - resuelto_fecha)
        stage1_summary[f"cumplimiento_{date_column}"] = float(resuelto_fecha / pedido_fecha) if pedido_fecha > 0 else 1.0

    stage2_summary = {
        "peso_disponible_total": peso_disponible_total,
        "peso_real_total": peso_real_total,
        "peso_ideal_resuelto_total": peso_ideal_resuelto_total,
        "peso_ideal_pedido_total": peso_ideal_total,
        "sobrepeso_real_vs_ideal": float(peso_real_total - peso_ideal_resuelto_total),
        "sobrepeso_pct_macro": float(cumplimiento_peso_macro - 1.0),
        "cumplimiento_peso_macro": cumplimiento_peso_macro,
        "holgura_peso_disponible": float(peso_disponible_total - peso_real_total),
        "tallos_disponibles_netos": tallos_disponibles_netos,
        "tallos_usados_netos": tallos_usados_netos,
        "tallos_restantes_netos": float(tallos_disponibles_netos - tallos_usados_netos),
        "tallos_brutos_usados": float(gross_matrix.sum()),
        "mallas_totales": float(mallas_matrix.sum()),
        "mallas_totales_redondeadas": float(excel_round(mallas_matrix.sum(), 0)),
        "tallos_brutos_redondeados": float(excel_round(gross_matrix.sum(), 0)),
        "peso_promedio_tallo_objetivo": peso_promedio_tallo_objetivo,
        "peso_promedio_tallo_disponible": peso_promedio_tallo_disponible,
        "peso_promedio_tallo_real": peso_promedio_tallo_real,
        "pedidos_en_objetivo": float(np.sum(within_objective & (fulfilled_bunches > 0))),
        "pedidos_fuera_objetivo": float(np.sum((fulfilled_bunches > 0) & ~within_objective)),
        "pedidos_sin_resolver": float(np.sum(fulfilled_bunches == 0)),
    }

    solver_meta = {
        "status": "Optimal"
        if all(
            status == "Optimal"
            for status in (
                stage1_status,
                stage1_status_macro,
                stage1_status_recipe,
                stage1_status_overweight,
                stage1_status_ideal,
                stage1_status_stems,
                stage2_status_pref,
                stage2_status_macro,
                stage2_status_balance,
                stage2_status_recipe,
                stage2_status_overweight,
                stage2_status_ideal,
                stage2_status_stems,
                stage2_status_extra,
                stage2_status_total,
                stage2_status_weight,
            )
        )
        else "Feasible con time limit",
        "stage1_macro_violation_opt": stage1_macro_violation_opt,
        "stage1_macro_target_deviation_opt": stage1_macro_target_deviation_opt,
        "stage1_recipe_gap_opt": stage1_recipe_gap_opt,
        "stage1_overweight_opt": stage1_overweight_opt,
        "stage1_ideal_deviation_opt": stage1_ideal_deviation_opt,
        "fulfilled_ideal_opt": fulfilled_ideal_opt,
        "macro_violation_opt": macro_violation_opt,
        "macro_target_deviation_opt": macro_target_deviation_opt if active_order_positions else 0.0,
        "balance_overweight_opt": balance_overweight_opt,
        "recipe_gap_opt": recipe_gap_opt,
        "preferred_violation_opt": preferred_violation_opt,
        "overweight_opt": overweight_opt,
        "ideal_deviation_opt": ideal_deviation_opt,
        "stem_violation_opt": stem_violation_opt,
        "extra_grades_opt": extra_grades_opt,
        "total_grades_opt": total_grades_opt,
        "stage1_status": stage1_status,
        "stage1_status_macro": stage1_status_macro,
        "stage1_status_recipe": stage1_status_recipe,
        "stage1_status_overweight": stage1_status_overweight,
        "stage1_status_ideal": stage1_status_ideal,
        "stage1_status_stems": stage1_status_stems,
        "stage1_stem_violation_opt": stage1_stem_violation_opt,
        "stage2_status_pref": stage2_status_pref,
        "stage2_status_macro": stage2_status_macro,
        "stage2_status_balance": stage2_status_balance,
        "stage2_status_recipe": stage2_status_recipe,
        "stage2_status_overweight": stage2_status_overweight,
        "stage2_status_ideal": stage2_status_ideal,
        "stage2_status_stems": stage2_status_stems,
        "stage2_status_extra": stage2_status_extra,
        "stage2_status_total": stage2_status_total,
        "stage2_status_weight": stage2_status_weight,
        "stage1_time_limit_seconds": STAGE1_TIME_LIMIT_SECONDS,
        "stage2_time_limit_seconds": STAGE2_TIME_LIMIT_SECONDS,
        "active_orders": len(active_order_positions),
        "active_grades": len(active_grade_positions),
    }
    for date_column in DATE_COLUMNS:
        solver_meta[f"stage1_priority_status_{date_column}"] = stage1_priority_statuses.get(date_column, "No aplica")
        solver_meta[f"stage1_priority_opt_{date_column}"] = stage1_priority_optima.get(date_column, 0.0)

    stage1_result = Stage1Result(
        orders=order_summary[
            [
                SKU_COLUMN,
                "pedido_total",
                "pedido_resuelto",
                "ajuste_bunches",
                "cumplimiento_bunches",
                "peso_ideal_pedido",
                "peso_ideal_resuelto",
            ]
        ].copy(),
        summary=stage1_summary,
    )
    stage2_result = Stage2Result(
        orders=order_summary,
        net_tallos=net_tallos,
        gross_tallos=gross_tallos,
        mallas_raw=mallas_raw,
        mallas_display=mallas_display,
        availability=availability_summary,
        summary=stage2_summary,
        solver_meta=solver_meta,
    )
    return PipelineResult(stage1=stage1_result, stage2=stage2_result)
